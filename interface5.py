import os
import csv
import hashlib
import threading
import time
import queue
from datetime import datetime, timedelta


# === Tkinter (Interface Gráfica) ===
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkinter import scrolledtext
from PIL import Image, ImageTk

from collections import Counter

# === Excel com openpyxl ===
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from collections import defaultdict

# === PDF com ReportLab ===
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Configurações de login
USUARIOS = {
    "000": {
        "senha": "c6f057b86584942e415435ffb1fa93d4",  # hash md5 de "000"
        "nome": "Administrador",
        "tipo": "bibliotecario"
    },
    "Ana Paula": {
        "senha": "202cb962ac59075b964b07152d234b70",  # hash md5 de "123"
        "nome": "Ana Paula",
        "tipo": "bibliotecario"
    }
}

# Dados de alunos (simulando banco de dados)
# FUTURO: integrar com empréstimos para mostrar histórico individual
ALUNOS = {
    "000": {
        "turma": "0",
        "nome": "João Silva",
        "senha": "202cb962ac59075b964b07152d234b70"  # hash md5 de "123"
    },
    "67890": {
        "turma": "8B",
        "nome": "Maria Oliveira",
        "senha": "250cf8b51c773f3f8dc8b4be867a9a02"  # hash md5 de "456"
    }
}







class LoginApp:
    def __init__(self, janela):
        self.janela = janela
        self.janela.title("Sistema de Biblioteca - Login")
        self.janela.attributes('-topmost', True)
        self.janela.bind("<Configure>", self._redimensionar_fundo)
        self._centralizar_janela()
        self._criar_interface_login_inicial()

    def _centralizar_janela(self):
        largura = 1000
        altura = 600
        
        largura_tela = self.janela.winfo_screenwidth()
        altura_tela = self.janela.winfo_screenheight()
        
        x = (largura_tela - largura) // 2
        y = (altura_tela - altura) // 2
        
        self.janela.geometry(f"{largura}x{altura}+{x}+{y}")
        self._carregar_imagem_fundo()

    def _carregar_imagem_fundo(self):
        try:
            self.img_original = Image.open("escola.png")
        except:
            self.img_original = None

        self.canvas_fundo = tk.Canvas(self.janela, highlightthickness=0)
        self.canvas_fundo.pack(fill="both", expand=True)

        self.bg_img = None
        if self.img_original:
            self._atualizar_imagem_fundo()

    def _atualizar_imagem_fundo(self):
        largura = self.janela.winfo_width()
        altura = self.janela.winfo_height()
        imagem_redimensionada = self.img_original.resize((largura, altura), Image.LANCZOS)
        self.bg_img = ImageTk.PhotoImage(imagem_redimensionada)
        self.canvas_fundo.create_image(0, 0, image=self.bg_img, anchor="nw")

    def _redimensionar_fundo(self, event):
        if self.img_original:
            self._atualizar_imagem_fundo()

    def _criar_interface_login_inicial(self):
        """Tela inicial com opções de login para aluno ou bibliotecário"""
        self.limpar_interface()
        
        # Frame semi-transparente sobre a imagem
        self.interface_frame = tk.Frame(self.canvas_fundo, bg="#ffffff", bd=2)
        self.interface_frame.place(relx=0.5, rely=0.5, anchor="center", width=400, height=300)

        try:
            logo_img = Image.open("logo.jpg")
            logo_img = logo_img.resize((80, 80), Image.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(logo_img)
            tk.Label(self.interface_frame, image=self.logo_img, bg="#ffffff").pack(pady=10)
        except:
            pass



        tk.Label(
            self.interface_frame,
            text="SISTEMA BIBLIOTECÁRIO",
            bg="#ffffff",
            fg="#2c3e50",
            font=("Arial", 20, "bold")
        ).pack(pady=(0, 20))

        # Botão de login para aluno
        btn_aluno = tk.Button(
            self.interface_frame,
            text="👩‍🎓 Acesso do Aluno",
            bg="#3498db",
            fg="white",
            font=("Arial", 12, "bold"),
            width=20,
            command=self._tela_login_aluno
        )
        btn_aluno.pack(pady=10)

        # Botão de login para bibliotecário
        btn_bibliotecario = tk.Button(
            self.interface_frame,
            text="👨‍💼 Acesso do Bibliotecário",
            bg="#2ecc71",
            fg="white",
            font=("Arial", 12, "bold"),
            width=20,
            command=self._tela_login_bibliotecario
        )
        btn_bibliotecario.pack(pady=10)

                # Rodapé
        footer = tk.Frame(self.interface_frame, bg="#ffffff")
        footer.pack(fill="x", pady=(10, 0))


        tk.Label(
            footer,
            text="v1.4.2",
            bg="#ffffff",
            fg="#999999",
            font=("Arial", 9)
        ).pack(side="right", padx=10)

    def _tela_login_aluno(self):
        """Tela de login específica para alunos"""
        self.limpar_interface()
        
        # Frame semi-transparente sobre a imagem
        self.interface_frame = tk.Frame(self.canvas_fundo, bg="#ffffff", bd=2)
        self.interface_frame.place(relx=0.525, rely=0.1, relwidth=0.4, relheight=0.8)


        try:
            logo_img = Image.open("logo.jpg")
            logo_img = logo_img.resize((60, 60), Image.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(logo_img)
            tk.Label(self.interface_frame, image=self.logo_img, bg="#ffffff").pack(pady=10)
        except:
            pass

        tk.Label(
            self.interface_frame,
            text="LOGIN DO ALUNO",
            bg="#ffffff",
            fg="#2c3e50",
            font=("Arial", 18, "bold")
        ).pack(pady=(10, 20))

        form_frame = tk.Frame(self.interface_frame, bg="#ffffff")
        form_frame.pack(padx=30, pady=10, fill="both", expand=True)

        # Campo Matrícula
        tk.Label(form_frame, text="Matrícula", bg="#ffffff", fg="#555555", anchor="w").pack(fill="x")
        self.matricula_entry = ttk.Entry(form_frame, font=("Arial", 12))
        self.matricula_entry.pack(fill="x", pady=(0, 10), ipady=8)
        self.matricula_entry.focus_set()

        # Campo Nome
        tk.Label(form_frame, text="nome", bg="#ffffff", fg="#555555", anchor="w").pack(fill="x")
        self.nome_entry = ttk.Entry(form_frame, font=("Arial", 12))
        self.nome_entry.pack(fill="x", pady=(0, 10), ipady=8)

        # Campo Senha
        tk.Label(form_frame, text="Senha", bg="#ffffff", fg="#555555", anchor="w", font=("Arial", 11)).pack(fill="x")

        senha_frame = tk.Frame(form_frame, bg="#ffffff")
        senha_frame.pack(fill="x", pady=(0, 20))

        self.senha_entry = ttk.Entry(senha_frame, font=("Arial", 12), show="•")
        self.senha_entry.grid(row=0, column=0, ipady=8, sticky="ew")

        # Estado de visibilidade da senha
        self.mostrar_senha = False

        def alternar_senha():
            self.mostrar_senha = not self.mostrar_senha
            self.senha_entry.config(show="" if self.mostrar_senha else "•")
            olho_btn.config(text="🙈" if self.mostrar_senha else "👁️")

        # Botão de olho mágico 👁️
        olho_btn = tk.Button(senha_frame, text="👁️", bg="#ffffff", bd=0, font=("Arial", 12),
                            command=alternar_senha, cursor="hand2", activebackground="#ffffff")
        olho_btn.grid(row=0, column=1, sticky="e")

        self.senha_entry.bind("<Return>", lambda event: self._fazer_login_aluno())

        # Botão de login
        login_btn = tk.Button(
            form_frame,
            text="ENTRAR",
            bg="#34db34",
            fg="white",
            font=("Arial", 12, "bold"),
            borderwidth=0,
            cursor="hand2",
            command=self._fazer_login_aluno
        )
        login_btn.pack(fill="x", pady=(2, 2))  # Espaçamento reduzido

        login_btn.bind("<Enter>", lambda e: login_btn.config(bg="#3cec3c"))
        login_btn.bind("<Leave>", lambda e: login_btn.config(bg="#34db34"))

        # Botão de voltar (também dentro do form_frame)
        voltar_btn = tk.Button(
            form_frame,  # <- Alterado para o mesmo frame
            text="⬅ Voltar",
            bg="#7f8c8d",
            fg="white",
            font=("Arial", 10),
            command=self._criar_interface_login_inicial
        )
        voltar_btn.pack(fill="x", pady=(2, 2))  # Espaço pequeno entre eles


    def _tela_login_bibliotecario(self):
        """Tela de login específica para bibliotecários"""
        self.limpar_interface()
        
        # Frame semi-transparente sobre a imagem
        self.interface_frame = tk.Frame(self.canvas_fundo, bg="#ffffff", bd=2)
        self.interface_frame.place(relx=0.525, rely=0.15, relwidth=0.4, relheight=0.7)

        try:
            logo_img = Image.open("logo.jpg")
            logo_img = logo_img.resize((70, 70), Image.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(logo_img)
            tk.Label(self.interface_frame, image=self.logo_img, bg="#ffffff").pack(pady=10)
        except:
            pass

        tk.Label(
            self.interface_frame,
            text="LOGIN DO BIBLIOTECÁRIO",
            bg="#ffffff",
            fg="#2c3e50",
            font=("Arial", 18, "bold")
        ).pack(pady=(10, 20))

        form_frame = tk.Frame(self.interface_frame, bg="#ffffff")
        form_frame.pack(padx=30, pady=10, fill="both", expand=True)

        # Campo Usuário
        tk.Label(form_frame, text="Usuário", bg="#ffffff", fg="#555555", anchor="w").pack(fill="x")
        self.usuario_entry = ttk.Entry(form_frame, font=("Arial", 12))
        self.usuario_entry.pack(fill="x", pady=(0, 10), ipady=8)
        self.usuario_entry.focus_set()

        # Campo Senha
        tk.Label(form_frame, text="Senha", bg="#ffffff", fg="#555555", anchor="w", font=("Arial", 11)).pack(fill="x")

        senha_frame = tk.Frame(form_frame, bg="#ffffff")
        senha_frame.pack(fill="x", pady=(0, 20))

        self.senha_entry = ttk.Entry(senha_frame, font=("Arial", 12), show="•")
        self.senha_entry.grid(row=0, column=0, ipady=8, sticky="ew")

        # Estado de visibilidade da senha
        self.mostrar_senha = False

        def alternar_senha():
            self.mostrar_senha = not self.mostrar_senha
            self.senha_entry.config(show="" if self.mostrar_senha else "•")
            olho_btn.config(text="🙈" if self.mostrar_senha else "👁️")

        # Botão de olho mágico 👁️
        olho_btn = tk.Button(senha_frame, text="👁️", bg="#ffffff", bd=0, font=("Arial", 12),
                            command=alternar_senha, cursor="hand2", activebackground="#ffffff")
        olho_btn.grid(row=0, column=1, sticky="e")

        self.senha_entry.bind("<Return>", lambda event: self._fazer_login_bibliotecario())

        # Botão de login
        login_btn = tk.Button(
            form_frame,
            text="ENTRAR",
            bg="#34db34",
            fg="white",
            font=("Arial", 12, "bold"),
            borderwidth=0,
            cursor="hand2",
            command=self._fazer_login_bibliotecario
        )
        login_btn.pack(fill="x", pady=(2, 2))  # Espaçamento reduzido

        login_btn.bind("<Enter>", lambda e: login_btn.config(bg="#3cec3c"))
        login_btn.bind("<Leave>", lambda e: login_btn.config(bg="#34db34"))

        # Botão de voltar (também no mesmo frame)
        voltar_btn = tk.Button(
            form_frame,  # <- aqui foi alterado
            text="⬅ Voltar",
            bg="#7f8c8d",
            fg="white",
            font=("Arial", 10),
            command=self._criar_interface_login_inicial
        )
        voltar_btn.pack(fill="x", pady=(2, 2))  # Espaço pequeno entre eles




    def limpar_interface(self):
        """Remove todos os widgets da interface"""
        for widget in self.canvas_fundo.winfo_children():
            if widget not in [self.canvas_fundo]:
                widget.destroy()

    def _fazer_login_aluno(self):
        """Verifica as credenciais do aluno e redireciona para a página principal"""
        try:
            matricula = self.matricula_entry.get().strip()
            senha = self.senha_entry.get().strip()
            
            if not matricula or not senha:
                messagebox.showerror("Erro", "Preencha todos os campos!")
                return
            
            # Verificar formato da matrícula (opcional)
            if not matricula.isdigit():
                messagebox.showerror("Erro", "Matrícula deve conter apenas números!")
                self.matricula_entry.delete(0, tk.END)
                return

            # Carregar planilha de alunos
            try:
                wb = openpyxl.load_workbook("alunos.xlsx")
                ws = wb.active
            except FileNotFoundError:
                messagebox.showerror("Erro", "Arquivo 'alunos.xlsx' não encontrado!")
                return
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível abrir a planilha: {str(e)}")
                return

            encontrado = False
            
            # Procurar aluno na planilha
            for row in ws.iter_rows(min_row=2):  # Ignora cabeçalho
                try:
                    # Colunas: 0=Matrícula, 1=Nome, 2=Turma, 3=Senha
                    matricula_plan = str(row[0].value).strip() if row[0].value else ""
                    nome_plan = str(row[1].value).strip() if row[1].value else ""
                    senha_plan = str(row[3].value).strip() if row[3].value else ""
                    
                    if matricula_plan == matricula:
                        encontrado = True
                        
                        # Verificar senha em texto puro
                        if senha_plan == senha:
                            messagebox.showinfo(
                                "Bem-vindo", 
                                f"Login realizado com sucesso!\n\nBem-vindo(a), {nome_plan}"
                            )
                            self.janela.destroy()

                            ALUNOS[matricula] = {
                                "nome": nome_plan,
                                "turma": str(row[2].value).strip() if row[2].value else "Turma não informada"
                            }

                            iniciar_sistema_aluno(matricula)


                           
                        else:
                            messagebox.showerror("Erro", "Senha incorreta!")
                            self.senha_entry.delete(0, tk.END)
                            return
                            
                except IndexError:
                    messagebox.showerror("Erro", "Planilha com formato inválido!")
                    return
                except Exception as e:
                    import traceback
                    traceback.print_exc()  # <<< Mostra o erro real no console
                    messagebox.showerror("Erro", f"Erro ao ler dados: {str(e)}")
                    return



            if not encontrado:
                messagebox.showerror(
                    "Aluno não cadastrado", 
                    "Matrícula não encontrada!\n\n"
                    "Solicite cadastro ao responsável."
                )
                self.matricula_entry.delete(0, tk.END)
                self.senha_entry.delete(0, tk.END)

        except Exception as e:
            messagebox.showerror("Erro Crítico", f"Falha inesperada: {str(e)}\n\nPor favor, reinicie o sistema.")
            # Para debug - mostra o erro completo no console
            import traceback
            traceback.print_exc()

    def _fazer_login_bibliotecario(self):
        """Verifica as credenciais do bibliotecário e redireciona para o sistema principal"""
        usuario = self.usuario_entry.get()
        senha = self.senha_entry.get()
        
        if not usuario or not senha:
            messagebox.showerror("Erro", "Preencha todos os campos!")
            return
        
        senha_hash = hashlib.md5(senha.encode()).hexdigest()
        
        if usuario in USUARIOS and USUARIOS[usuario]["senha"] == senha_hash:
            messagebox.showinfo(
                "Bem-vindo", 
                f"Login realizado com sucesso!\n\nBem-vindo, {USUARIOS[usuario]['nome']}"
            )
            self.janela.destroy()
            iniciar_sistema_principal()
        else:
            messagebox.showerror(
                "Erro", 
                "Credenciais inválidas!\n\nVerifique seu usuário e senha."
            )
            self.senha_entry.delete(0, tk.END)


    

# Configurações globais do sistema
ARQUIVO_PLANILHA = "LIVROS LITERÁRIOS-  BILBIOTECA ANA ZÉLIA DE MORAIS LARA.xlsx"
ARQUIVO_DOACAO = "livros.xlsx"
ARQUIVO_EMPRESTIMOS = "Emprestimos.xlsx"
CAMPOS = ["PRATELEIRA", "NUMERAÇÃO", "LIVROS", "AUTOR(ES)", "EDITORAS", "CLASSIFICAÇÃO/TEMA", "FAIXA ETÁRIA"]
CAMPOS_EMPRESTIMO = ["Código", "Título", "Autor", "Aluno", "Turma", "Prazo (dias)", "Data do Empréstimo"]

# Cache otimizado
MEMORY_CACHE = {
    'livros': [],
    'livros_doacao': [],
    'emprestimos': [],
    'generos': set(),
    'turmas': set(),
    'last_update': 0,
    'ready': False
}

# Fila para comunicação entre threads
DATA_QUEUE = queue.Queue()

def criar_planilha_if_not_exists():
    if not os.path.exists(ARQUIVO_PLANILHA):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(CAMPOS)
        wb.save(ARQUIVO_PLANILHA)
    
    if not os.path.exists(ARQUIVO_DOACAO):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(CAMPOS)
        wb.save(ARQUIVO_DOACAO)
    
    if not os.path.exists(ARQUIVO_EMPRESTIMOS):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(CAMPOS_EMPRESTIMO)
        wb.save(ARQUIVO_EMPRESTIMOS)

def carregar_dados_livros():
    try:
        wb = openpyxl.load_workbook(ARQUIVO_PLANILHA, read_only=True, data_only=True)
        ws = wb.active
        generos = set()
        dados = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(row):
                dados.append(row)
                if len(row) > 5 and row[5]:
                    generos.add(str(row[5]))
        
        MEMORY_CACHE['generos'] = generos
        return dados
    except Exception as e:
        print(f"Erro ao carregar livros: {str(e)}")
        return []

def carregar_dados_livros_doacao():
    try:
        wb = openpyxl.load_workbook(ARQUIVO_DOACAO, read_only=True, data_only=True)
        ws = wb.active
        dados = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(row):
                dados.append(row)
        
        return dados
    except Exception as e:
        print(f"Erro ao carregar livros doação: {str(e)}")
        return []

def carregar_dados_emprestimos():
    try:
        wb = openpyxl.load_workbook(ARQUIVO_EMPRESTIMOS, read_only=True, data_only=True)
        ws = wb.active
        turmas = set()
        dados = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(row):
                dados.append(row)
                if len(row) > 4 and row[4]:
                    turmas.add(row[4])
        
        MEMORY_CACHE['turmas'] = turmas
        return dados
    except Exception as e:
        print(f"Erro ao carregar empréstimos: {str(e)}")
        return []

def carregar_dados_em_segundo_plano():
    global MEMORY_CACHE

    dados_livros = carregar_dados_livros()
    dados_livros_doacao = carregar_dados_livros_doacao()
    dados_emprestimos = carregar_dados_emprestimos()

    MEMORY_CACHE['livros'] = dados_livros
    MEMORY_CACHE['livros_doacao'] = dados_livros_doacao
    MEMORY_CACHE['emprestimos'] = dados_emprestimos
    MEMORY_CACHE['last_update'] = time.time()
    MEMORY_CACHE['ready'] = True

    DATA_QUEUE.put(('dados_carregados', len(dados_livros), len(dados_livros_doacao), len(dados_emprestimos)))


def iniciar_sistema_principal():
    criar_planilha_if_not_exists()
    
    janela = tk.Tk()
    janela.title("Sistema de Gerenciamento de Livros")
    janela.geometry("1100x650")
    janela.configure(bg="#2c2c2c")

    # === Controle de inatividade ===
    janela.last_activity = time.time()
    
    def reset_inactivity_timer(event=None):
        janela.last_activity = time.time()
    
    def check_inactivity():
        if time.time() - janela.last_activity > 1800:  # 30 minutos
            if messagebox.askyesno("Sessão Expirada", 
                                  "Sua sessão expirou por inatividade. Deseja continuar?"):
                reset_inactivity_timer()
            else:
                janela.destroy()
                root = tk.Tk()
                app = LoginApp(root)
                root.mainloop()
        else:
            janela.after(60000, check_inactivity)  # Verificar a cada minuto
    
    # Registrar eventos para resetar o timer
    janela.bind("<Key>", reset_inactivity_timer)
    janela.bind("<Button>", reset_inactivity_timer)
    janela.after(60000, check_inactivity)  # Iniciar verificação
    
    # Componentes da interface
    topo = tk.Frame(janela, bg="#34db34", height=70)
    topo.pack(side="top", fill="x")
    
    titulo = tk.Label(
        topo,
        text="📖 BIBLIOTECA SMART | ANA ZÉLIA DE MORAIS LARA",
        bg="#34db34",
        fg="white",
        font=("Segoe UI", 18, "bold")
    )
    titulo.pack(pady=10, expand=True)
    
    menu = tk.Frame(janela, bg="#1c1c1c", width=240)
    menu.pack(side="left", fill="y")
    
    # Criar container para botões do menu
    menu_content = tk.Frame(menu, bg="#1c1c1c")
    menu_content.pack(side="top", fill="both", expand=True)
    
    # Criar frame para botão de logout no rodapé
    menu_bottom = tk.Frame(menu, bg="#1c1c1c")
    menu_bottom.pack(side="bottom", fill="x")
    
    conteudo = tk.Frame(janela, bg="white")
    conteudo.pack(expand=True, fill="both")
    
    carregando_frame = tk.Frame(conteudo, bg="white")
    carregando_frame.pack(expand=True, fill="both")
    tk.Label(carregando_frame, text="Carregando dados, aguarde...", font=("Arial", 14), bg="white").pack(expand=True)
    
    estilo_botao = {
        "bg": "#333333",
        "fg": "white",
        "font": ("Segoe UI", 10),
        "relief": "flat",
        "width": 30,
        "anchor": "w",
        "padx": 10,
        "pady": 5
    }
    
    def limpar_conteudo():
        for widget in conteudo.winfo_children():
            widget.destroy()
    
    def novo_livro():
        limpar_conteudo()
        entradas = []

        tk.Label(
            conteudo,
            text="Cadastro de Novo Livro",
            font=("Segoe UI", 14, "bold"),
            bg="white"
        ).pack(pady=15)
        
        for campo in CAMPOS:
            tk.Label(conteudo, text=campo, font=("Segoe UI", 12), bg="white").pack(pady=4)
            entry = tk.Entry(conteudo, width=55, font=("Segoe UI", 12))
            entry.pack(pady=2)
            entradas.append(entry)

        # Frame para os botões
        botoes_frame = tk.Frame(conteudo, bg="white")
        botoes_frame.pack(pady=10)

        def salvar():
            dados = [e.get() for e in entradas]
            if not all(dados):
                messagebox.showerror("Erro", "Preencha todos os campos.")
                return
                
            try:
                wb = openpyxl.load_workbook(ARQUIVO_PLANILHA)
                ws = wb.active
                ws.append(dados)
                wb.save(ARQUIVO_PLANILHA)
                
                threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
                messagebox.showinfo("Sucesso", "Livro cadastrado com sucesso!")
                exibir_livros()
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao salvar: {str(e)}")

        def adicionar_em_massa():
            # Verifica se todos os campos estão preenchidos
            dados_base = [e.get() for e in entradas]
            if not all(dados_base):
                messagebox.showerror("Erro", "Preencha todos os campos primeiro.")
                return

            # Pede a quantidade de livros
            quantidade = simpledialog.askinteger(
                "Adicionar em Massa",
                "Quantos livros deseja cadastrar?",
                parent=janela,
                minvalue=2,
                maxvalue=100
            )
            
            if not quantidade:
                return

            # Pega o código base (assumindo que o código está no segundo campo - NUMERAÇÃO)
            codigo_base = dados_base[1]  # Índice 1 para NUMERAÇÃO
            try:
                # Extrai a parte numérica do código (assumindo formato XXXXX-X)
                parte_principal, digito = codigo_base.split('-')
                numero_base = int(parte_principal)
                digito_base = int(digito)
                
                wb = openpyxl.load_workbook(ARQUIVO_PLANILHA)
                ws = wb.active
                
                # Progress bar para feedback visual
                progresso = ttk.Progressbar(conteudo, orient="horizontal", 
                                        length=300, mode="determinate")
                progresso.pack(pady=10)
                janela.update()  # Atualiza a interface
                
                # Cadastra cada livro com código sequencial
                for i in range(quantidade):
                    novo_codigo = f"{numero_base + i}-{(digito_base + i) % 10}"
                    dados = dados_base.copy()
                    dados[1] = novo_codigo  # Atualiza o campo NUMERAÇÃO
                    ws.append(dados)
                    
                    # Atualiza a barra de progresso
                    progresso['value'] = (i + 1) / quantidade * 100
                    janela.update()
                
                wb.save(ARQUIVO_PLANILHA)
                progresso.destroy()
                
                threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
                messagebox.showinfo("Sucesso", f"{quantidade} livros cadastrados com sucesso!")
                exibir_livros()
                
            except ValueError:
                messagebox.showerror("Erro", "Formato de código inválido. Use o formato 00000-0")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao cadastrar em massa: {str(e)}")
                if 'progresso' in locals():
                    progresso.destroy()

        # Botão Salvar único
        tk.Button(botoes_frame, text="Salvar Livro", command=salvar, 
                bg="#4CAF50", fg="white", font=("Segoe UI", 12)).pack(side="left", padx=10)

        # Botão Adicionar em Massa
        tk.Button(botoes_frame, text="Adicionar em Massa", command=adicionar_em_massa,
                bg="#2196F3", fg="white", font=("Segoe UI", 12)).pack(side="left", padx=10)
    

    def novo_livro_doacao():
        limpar_conteudo()
        entradas = []

        tk.Label(
            conteudo,
            text="Cadastro de Novo Livro da Doação",
            font=("Segoe UI", 14, "bold"),
            bg="white"
        ).pack(pady=15)
        
        for campo in CAMPOS:
            tk.Label(conteudo, text=campo, font=("Segoe UI", 12), bg="white").pack(pady=4)
            entry = tk.Entry(conteudo, width=55, font=("Segoe UI", 12))
            entry.pack(pady=2)
            entradas.append(entry)

        # Frame para os botões
        botoes_frame = tk.Frame(conteudo, bg="white")
        botoes_frame.pack(pady=10)

        def salvar():
            dados = [e.get() for e in entradas]
            if not all(dados):
                messagebox.showerror("Erro", "Preencha todos os campos.")
                return
                
            try:
                wb = openpyxl.load_workbook(ARQUIVO_DOACAO)
                ws = wb.active
                ws.append(dados)
                wb.save(ARQUIVO_DOACAO)
                
                threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
                messagebox.showinfo("Sucesso", "Livro de doação cadastrado com sucesso!")
                exibir_livros_doacao()
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao salvar: {str(e)}")

        def adicionar_em_massa():
            # Verifica se todos os campos estão preenchidos
            dados_base = [e.get() for e in entradas]
            if not all(dados_base):
                messagebox.showerror("Erro", "Preencha todos os campos primeiro.")
                return

            # Pede a quantidade de livros
            quantidade = simpledialog.askinteger(
                "Adicionar em Massa",
                "Quantos livros deseja cadastrar?",
                parent=janela,
                minvalue=2,
                maxvalue=100
            )
            
            if not quantidade:
                return

            # Pega o código base (assumindo que o código está no segundo campo)
            codigo_base = dados_base[1]  # Índice 1 para NUMERAÇÃO
            try:
                # Extrai a parte numérica do código (assumindo formato XXXXX-X)
                parte_principal, digito = codigo_base.split('-')
                numero_base = int(parte_principal)
                digito_base = int(digito)
                
                wb = openpyxl.load_workbook(ARQUIVO_DOACAO)
                ws = wb.active
                
                # Cadastra cada livro com código sequencial
                for i in range(quantidade):
                    novo_codigo = f"{numero_base + i}-{(digito_base + i) % 10}"
                    dados = dados_base.copy()
                    dados[1] = novo_codigo  # Atualiza o campo NUMERAÇÃO
                    ws.append(dados)
                
                wb.save(ARQUIVO_DOACAO)
                
                threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
                messagebox.showinfo("Sucesso", f"{quantidade} livros cadastrados com sucesso!")
                exibir_livros_doacao()
                
            except ValueError:
                messagebox.showerror("Erro", "Formato de código inválido. Use o formato 00000-0")
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao cadastrar em massa: {str(e)}")

        # Botão Salvar único
        tk.Button(botoes_frame, text="Salvar Livro", command=salvar, 
                bg="#4CAF50", fg="white", font=("Segoe UI", 12)).pack(side="left", padx=10)

        # Botão Adicionar em Massa
        tk.Button(botoes_frame, text="Adicionar em Massa", command=adicionar_em_massa,
                bg="#2196F3", fg="white", font=("Segoe UI", 12)).pack(side="left", padx=10)    


    def cadastrar_aluno():
        limpar_conteudo()

        tk.Label(
            conteudo,
            text="Cadastro de Novo Aluno",
            font=("Segoe UI", 14, "bold"),
            bg="white"
        ).pack(pady=15)

        entradas = {}

        campos = ["Matrícula", "Nome", "Turma", "Senha"]
        for campo in campos:
            tk.Label(conteudo, text=campo, font=("Segoe UI", 12), bg="white").pack(pady=4)
            entry = tk.Entry(conteudo, width=55, font=("Segoe UI", 12), show="•" if campo == "Senha" else "")
            entry.pack(pady=2)
            entradas[campo.lower()] = entry

        # Botões
        botoes_frame = tk.Frame(conteudo, bg="white")
        botoes_frame.pack(pady=15)

        def salvar_aluno():
            matricula = entradas["matrícula"].get().strip()
            nome = entradas["nome"].get().strip()
            turma = entradas["turma"].get().strip()
            senha = entradas["senha"].get().strip()

            if not all([matricula, nome, turma, senha]):
                messagebox.showerror("Erro", "Todos os campos devem ser preenchidos!")
                return

            arquivo = "alunos.xlsx"

            try:
                # Se o arquivo não existe, cria e adiciona cabeçalho
                if not os.path.exists(arquivo):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Alunos"
                    ws.append(["Matrícula", "Nome", "Turma", "Senha"])
                    wb.save(arquivo)

                # Abre o arquivo existente
                wb = openpyxl.load_workbook(arquivo)
                ws = wb.active

                # Adiciona os dados na próxima linha vazia
                ws.append([matricula, nome, turma, senha])
                wb.save(arquivo)

                messagebox.showinfo("Sucesso", "Aluno cadastrado com sucesso!")

                # Limpa os campos de entrada
                for entry in entradas.values():
                    entry.delete(0, tk.END)

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar cadastro: {str(e)}")



        tk.Button(botoes_frame, text="Salvar Aluno", command=salvar_aluno,
                bg="#4CAF50", fg="white", font=("Segoe UI", 12)).pack(side="left", padx=10)



    def atualizar_livro(index):
        if not MEMORY_CACHE['ready']:
            messagebox.showinfo("Aguarde", "Dados ainda não carregados completamente")
            return
            
        limpar_conteudo()
        dados_livros = MEMORY_CACHE['livros']
        
        if index >= len(dados_livros):
            messagebox.showerror("Erro", "Índice inválido")
            return
            
        livro_selecionado = dados_livros[index]
        entradas = []

        # Interface para edição (mantida igual)
        for i, campo in enumerate(CAMPOS):
            tk.Label(conteudo, text=campo, font=("Segoe UI", 10), bg="white").pack(pady=2)
            entry = tk.Entry(conteudo, width=50)
            entry.insert(0, livro_selecionado[i] if i < len(livro_selecionado) else "")
            entry.pack()
            entradas.append(entry)

        def salvar_atualizacao():
            novos_dados = [e.get() for e in entradas]
            if not all(novos_dados):
                messagebox.showerror("Erro", "Todos os campos devem ser preenchidos.")
                return
                
            try:
                wb = openpyxl.load_workbook(ARQUIVO_PLANILHA)
                ws = wb.active
                
                # Buscar a linha exata correspondente ao livro selecionado
                linha_encontrada = None
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    valores_linha = [cell.value for cell in row]
                    
                    # Verificar se é a linha correta comparando todos os valores
                    corresponde = True
                    for i in range(min(len(valores_linha), len(livro_selecionado))):
                        if str(valores_linha[i]) != str(livro_selecionado[i]):
                            corresponde = False
                            break
                    
                    if corresponde:
                        linha_encontrada = row_idx
                        break
                
                if linha_encontrada is None:
                    messagebox.showerror("Erro", "Livro não encontrado na planilha!")
                    return
                
                # Atualizar os dados na linha encontrada
                for i in range(len(CAMPOS)):
                    ws.cell(row=linha_encontrada, column=i+1).value = novos_dados[i]
                
                wb.save(ARQUIVO_PLANILHA)
                
                # Atualizar cache e interface
                threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
                messagebox.showinfo("Atualizado", "Livro atualizado com sucesso!")
                exibir_livros()
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao atualizar: {str(e)}")

        tk.Button(conteudo, 
                text="Salvar alterações", 
                command=salvar_atualizacao, 
                bg="#4CAF50", 
                fg="white").pack(pady=10)

    def atualizar_livro_doacao(index):
        if not MEMORY_CACHE['ready']:
            messagebox.showinfo("Aguarde", "Dados ainda não carregados completamente")
            return
            
        limpar_conteudo()
        dados_livros_doacao = MEMORY_CACHE['livros_doacao']
        
        if index >= len(dados_livros_doacao):
            messagebox.showerror("Erro", "Índice inválido")
            return
            
        livro_selecionado = dados_livros_doacao[index]
        entradas = []

        for i, campo in enumerate(CAMPOS):
            tk.Label(conteudo, text=campo, font=("Segoe UI", 10), bg="white").pack(pady=2)
            entry = tk.Entry(conteudo, width=50)
            entry.insert(0, livro_selecionado[i] if i < len(livro_selecionado) else "")
            entry.pack()
            entradas.append(entry)

        def salvar_atualizacao():
            novos_dados = [e.get() for e in entradas]
            if not all(novos_dados):
                messagebox.showerror("Erro", "Todos os campos devem ser preenchidos.")
                return
                
            try:
                wb = openpyxl.load_workbook(ARQUIVO_DOACAO)
                ws = wb.active
                
                # Encontrar a linha exata correspondente ao livro selecionado
                linha_encontrada = None
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    valores_linha = [cell.value for cell in row]
                    
                    # Verificar se é a linha que queremos atualizar
                    corresponde = True
                    for i in range(min(len(valores_linha), len(livro_selecionado))):
                        if str(valores_linha[i]) != str(livro_selecionado[i]):
                            corresponde = False
                            break
                    
                    if corresponde:
                        linha_encontrada = row_idx
                        break
                
                if linha_encontrada is None:
                    messagebox.showerror("Erro", "Livro não encontrado na planilha!")
                    return
                
                # Atualizar os dados na linha encontrada
                for i in range(len(CAMPOS)):
                    ws.cell(row=linha_encontrada, column=i+1).value = novos_dados[i]
                
                wb.save(ARQUIVO_DOACAO)
                
                threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
                messagebox.showinfo("Atualizado", "Livro de doação atualizado com sucesso!")
                exibir_livros_doacao()
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao atualizar: {str(e)}")

        tk.Button(conteudo, text="Salvar alterações", command=salvar_atualizacao, 
                bg="#4CAF50", fg="white").pack(pady=10)
    
    def excluir_livro(index):
        if not MEMORY_CACHE['ready']:
            messagebox.showinfo("Aguarde", "Dados ainda não carregados completamente")
            return
        
        senha_correta = "123" 
        senha_digitada = simpledialog.askstring("Senha", "Digite a senha para excluir:", show='*')
        
        if senha_digitada != senha_correta:
            messagebox.showerror("Erro", "Senha incorreta!")
            return
        
        if not messagebox.askyesno("Confirmar", "Deseja excluir este livro?"):
            return
            
        try:
            # Obter os dados do livro selecionado do cache
            livro_selecionado = MEMORY_CACHE['livros'][index]
            
            # Abrir a planilha existente
            wb = openpyxl.load_workbook(ARQUIVO_PLANILHA)
            ws = wb.active
            
            # Criar uma nova planilha em memória
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = ws.title
            
            # Copiar cabeçalhos
            for row in ws.iter_rows(max_row=1):
                values = [cell.value for cell in row]
                new_ws.append(values)
            
            # Copiar todas as linhas exceto a que será excluída
            for row in ws.iter_rows(min_row=2):
                valores_linha = [cell.value for cell in row]
                
                # Verificar se a linha atual corresponde ao livro selecionado
                corresponde = True
                for i in range(min(len(valores_linha), len(livro_selecionado))):
                    if str(valores_linha[i]) != str(livro_selecionado[i]):
                        corresponde = False
                        break
                
                if not corresponde:
                    new_ws.append(valores_linha)
            
            # Salvar a nova planilha
            new_wb.save(ARQUIVO_PLANILHA)
            
            # Atualizar o cache e a interface
            threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
            messagebox.showinfo("Sucesso", "Livro excluído com sucesso!")
            exibir_livros()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao excluir: {str(e)}")


    def excluir_livros_doacao(indices=None):
        """Função corrigida para apagar TODOS os itens selecionados sem falhas"""
        if not MEMORY_CACHE['ready']:
            messagebox.showinfo("Aguarde", "Dados ainda não carregados completamente")
            return

        # Verificação de senha
        senha_correta = "123"
        senha_digitada = simpledialog.askstring("Senha", "Digite a senha para excluir:", show='*')
        
        if senha_digitada != senha_correta:
            messagebox.showerror("Erro", "Senha incorreta!")
            return

        # Verificar e processar os índices
        if indices is None:
            return
        if isinstance(indices, int):
            indices = [indices]
        
        # Obter TODOS os livros atuais
        todos_livros = MEMORY_CACHE['livros_doacao']
        
        # Identificar os livros que SERÃO REMOVIDOS
        livros_para_remover = [todos_livros[i] for i in indices]
        
        # Confirmação FINAL
        if not messagebox.askyesno("Confirmar", f"Tem certeza que deseja excluir {len(livros_para_remover)} livro(s)?"):
            return

        try:
            # Abrir o arquivo Excel ORIGINAL
            wb = openpyxl.load_workbook(ARQUIVO_DOACAO)
            ws = wb.active
            
            # Obter TODAS as linhas do Excel
            todas_linhas_excel = [linha for linha in ws.iter_rows(values_only=True)]
            
            # Lista para as linhas que VÃO PERMANECER
            linhas_que_ficam = [todas_linhas_excel[0]]  # Cabeçalho
            
            # Filtro AVANÇADO para remoção
            for linha_excel in todas_linhas_excel[1:]:
                # Verificar se esta linha DEVE SER REMOVIDA
                deve_remover = False
                for livro_remover in livros_para_remover:
                    # Comparação PROFUNDA entre os campos
                    campos_iguais = True
                    for i in range(min(len(linha_excel), len(livro_remover))):
                        if str(linha_excel[i]).strip() != str(livro_remover[i]).strip():
                            campos_iguais = False
                            break
                    
                    if campos_iguais:
                        deve_remover = True
                        break
                
                if not deve_remover:
                    linhas_que_ficam.append(linha_excel)
            
            # Criar NOVO arquivo Excel apenas com as linhas que ficam
            novo_wb = openpyxl.Workbook()
            nova_ws = novo_wb.active
            
            for linha in linhas_que_ficam:
                nova_ws.append(linha)
            
            # Salvar SUBSTITUINDO o arquivo original
            novo_wb.save(ARQUIVO_DOACAO)
            
            # Atualização IMEDIATA do sistema
            MEMORY_CACHE['livros_doacao'] = linhas_que_ficam[1:]  # Exclui cabeçalho
            exibir_livros_doacao()  # Atualiza a interface AGORA
            
            messagebox.showinfo("Concluído", f"{len(livros_para_remover)} livro(s) removido(s) com sucesso!")
        
        except Exception as e:
            messagebox.showerror("Falha", f"Erro durante exclusão: {str(e)}")

    # Compatibilidade com código antigo
    excluir_livro_doacao = excluir_livros_doacao

    class Paginacao:
        def __init__(self, dados, page_size=100):
            self.dados = dados
            self.page_size = page_size
            self.current_page = 0
            self.total_pages = max(1, (len(dados) + page_size - 1) // page_size)
        
        def get_page(self, page_num=None):
            if page_num is None:
                page_num = self.current_page
            start = page_num * self.page_size
            end = start + self.page_size
            return self.dados[start:end]
        
        def next_page(self):
            if self.current_page < self.total_pages - 1:
                self.current_page += 1
            return self.get_page()
        
        def prev_page(self):
            if self.current_page > 0:
                self.current_page -= 1
            return self.get_page()
    
    def exibir_livros():
        limpar_conteudo()
        
        if not MEMORY_CACHE['ready']:
            carregando_frame = tk.Frame(conteudo, bg="white")
            carregando_frame.pack(expand=True, fill="both")
            tk.Label(carregando_frame, text="Carregando dados, aguarde...", font=("Arial", 14), bg="white").pack(expand=True)
            return
        
        dados_livros = MEMORY_CACHE['livros']
        paginacao = Paginacao(dados_livros)

        # Frame de filtro simplificado
        filtro_frame = tk.Frame(conteudo, bg="white")
        filtro_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(filtro_frame, text="🔍 Pesquisar (título, autor, editora, tema, etc.):", bg="white").pack(side="left")
        entrada_pesquisa = tk.Entry(filtro_frame, width=40)
        entrada_pesquisa.pack(side="left", padx=5)
        entrada_pesquisa.focus_set()  # Foco automático no campo de pesquisa

        # Botão de reiniciar/atualizar planilha
        def reiniciar_planilha():
            """Atualização rápida com feedback visual"""
            loading_label = tk.Label(conteudo, text="🔄 Atualizando...", font=("Arial", 12), bg="white")
            loading_label.pack(pady=10)
            janela.update()  # Atualiza a interface imediatamente
            
            def tarefa_atualizacao():
                try:
                    mod_time = os.path.getmtime(ARQUIVO_PLANILHA)
                    
                    if mod_time > MEMORY_CACHE['last_update']:
                        dados = carregar_dados_livros()
                        MEMORY_CACHE['livros'] = dados
                        MEMORY_CACHE['last_update'] = time.time()
                        
                        janela.after(0, lambda: [
                            loading_label.destroy(),
                            exibir_livros(),
                            messagebox.showinfo("Sucesso", "Planilha atualizada com sucesso!")
                        ])
                    else:
                        janela.after(0, lambda: [
                            loading_label.destroy(),
                            messagebox.showinfo("Info", "Nenhuma alteração encontrada na planilha.")
                        ])
                        
                except Exception as e:
                    janela.after(0, lambda: [
                        loading_label.destroy(),
                        messagebox.showerror("Erro", f"Falha ao atualizar:\n{str(e)}")
                    ])
            
            threading.Thread(target=tarefa_atualizacao, daemon=True).start()
        
        tk.Button(filtro_frame, text="🔄 Atualizar Planilha", command=reiniciar_planilha, 
                bg="#16E127", fg="white").pack(side="right", padx=5)

        # Tabela
        tree_frame = tk.Frame(conteudo, bg="white")
        tree_frame.pack(fill="both", expand=True)

        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")

        # Usar lista para colunas
        tree = ttk.Treeview(tree_frame, columns=["#"] + CAMPOS, show="headings", 
                        yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        tree.pack(fill="both", expand=True)
        scroll_y.config(command=tree.yview)
        scroll_x.config(command=tree.xview)

        tree.heading("#", text="Nº")
        tree.column("#", width=50, anchor="center")

        for campo in CAMPOS:
            tree.heading(campo, text=campo)
            tree.column(campo, anchor="center", width=150)

        filtered_data = []
        current_pagination = paginacao

        def filtrar_livros():
            nonlocal filtered_data, current_pagination
            
            termo = entrada_pesquisa.get().lower()
            
            filtered_data = []
            
            for row in dados_livros:
                if not row or len(row) < 2:
                    continue
                    
                # Verifica em todos os campos relevantes
                campos_relevantes = [
                    str(row[0]) if len(row) > 0 else "",  # PRATELEIRA
                    str(row[1]) if len(row) > 1 else "",  # NUMERAÇÃO
                    str(row[2]) if len(row) > 2 else "",  # LIVROS (título)
                    str(row[3]) if len(row) > 3 else "",  # AUTOR(ES)
                    str(row[4]) if len(row) > 4 else "",  # EDITORAS
                    str(row[5]) if len(row) > 5 else "",  # CLASSIFICAÇÃO/TEMA
                    str(row[6]) if len(row) > 6 else ""   # FAIXA ETÁRIA
                ]
                
                # Verifica se o termo de pesquisa está em qualquer campo relevante
                if termo and any(termo in campo.lower() for campo in campos_relevantes):
                    filtered_data.append(row)
                elif not termo:  # Se não há termo de pesquisa, mostra tudo
                    filtered_data.append(row)
            
            current_pagination = Paginacao(filtered_data)
            update_table()

        def update_table():
            nonlocal current_pagination
            
            tree.delete(*tree.get_children())
            page_data = current_pagination.get_page()
            
            for i, row in enumerate(page_data, start=current_pagination.current_page * current_pagination.page_size):
                tree.insert("", "end", iid=i, values=(i+1,) + row)
            
            label_pagina.config(text=f"Página {current_pagination.current_page+1}/{current_pagination.total_pages}")
            botao_anterior.config(state="normal" if current_pagination.current_page > 0 else "disabled")
            botao_proximo.config(state="normal" if current_pagination.current_page < current_pagination.total_pages - 1 else "disabled")

        tk.Button(filtro_frame, text="🔎 Filtrar", command=filtrar_livros, bg="#2196F3", fg="white").pack(side="left", padx=10)

        # Controles de paginação
        paginacao_frame = tk.Frame(conteudo, bg="white")
        paginacao_frame.pack(pady=5)

        botao_anterior = tk.Button(paginacao_frame, text="◀ Anterior", command=lambda: [current_pagination.prev_page(), update_table()])
        botao_anterior.pack(side="left", padx=5)

        label_pagina = tk.Label(paginacao_frame, text=f"Página 1/{paginacao.total_pages}", bg="white")
        label_pagina.pack(side="left", padx=5)

        botao_proximo = tk.Button(paginacao_frame, text="Próximo ▶", command=lambda: [current_pagination.next_page(), update_table()])
        botao_proximo.pack(side="left", padx=5)

        # Preencher tabela inicial
        filtered_data = dados_livros
        update_table()

        # Botões de ação
        def get_selected_index():
            selecionado = tree.selection()
            if not selecionado:
                messagebox.showerror("Erro", "Selecione um livro.")
                return None
            return int(selecionado[0])

        acoes = tk.Frame(conteudo, bg="white")
        acoes.pack(pady=10)
        tk.Button(acoes, text="✏️ Atualizar", command=lambda: atualizar_livro(get_selected_index()), bg="#4CAF50", fg="white").pack(side="left", padx=10)
        tk.Button(acoes, text="🗑️ Excluir", command=lambda: excluir_livro(get_selected_index()), bg="#f44336", fg="white").pack(side="left", padx=10)

 
        # Exportação
        def exportar_csv(dados, titulo="ACERVO - LIVROS LITERÁRIOS - BIBLIOTECA ANA ZÉLIA DE MORAIS LARA"):
            escolha = messagebox.askquestion(
                "Exportar", 
                "Deseja exportar a página atual ou todo o conteúdo?",
                icon='question', type='yesnocancel',
                detail="Selecione 'Sim' para página atual, 'Não' para todo conteúdo"
            )
            
            if escolha == 'cancel':
                return

            dados_exportar = current_pagination.get_page() if escolha == 'yes' else dados

            caminho = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Planilhas Excel", "*.xlsx")],
                title="Salvar como Excel",
                initialfile=f"acervo_livros_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            )

            if not caminho:
                return

            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Acervo"

                # Estilos
                negrito = Font(bold=True)
                centralizado = Alignment(horizontal="center", vertical="center")
                borda_fina = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Título
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CAMPOS))
                cell_titulo = ws.cell(row=1, column=1)
                cell_titulo.value = titulo
                cell_titulo.font = Font(size=14, bold=True)
                cell_titulo.alignment = centralizado

                # Cabeçalho
                for col, campo in enumerate(CAMPOS, start=1):
                    cell = ws.cell(row=3, column=col, value=campo)
                    cell.font = negrito
                    cell.alignment = centralizado
                    cell.border = borda_fina

                # Dados
                for i, linha in enumerate(dados_exportar, start=4):
                    for j, valor in enumerate(linha, start=1):
                        cell = ws.cell(row=i, column=j, value=valor)
                        cell.alignment = Alignment(horizontal="left")
                        cell.border = borda_fina

                # Ajustar largura das colunas automaticamente
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column  # número da coluna
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ajuste = max_length + 2
                    ws.column_dimensions[get_column_letter(column)].width = ajuste

                # Linha de rodapé
                ws.append([])
                ws.append([f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
                ws.append([f"Total de registros: {len(dados_exportar)}"])

                # Salvar
                wb.save(caminho)

                messagebox.showinfo(
                    "Exportado com sucesso",
                    f"Planilha Excel gerada com sucesso!\n\nLocal: {caminho}\nRegistros exportados: {len(dados_exportar)}\nTipo: {'Página atual' if escolha == 'yes' else 'Todo o conteúdo'}",
                    parent=janela
                )

            except Exception as e:
                messagebox.showerror("Erro na exportação", f"Erro ao exportar os dados:\n{str(e)}", parent=janela)
            
        def exportar_pdf(dados, titulo="ACERVO - LIVROS LITERÁRIOS - BIBLIOTECA ANA ZÉLIA DE MORAIS LARA"):
            orientacao = messagebox.askquestion("Orientação", 
                                            "Deseja página na vertical ou horizontal?", 
                                            icon='question', type='yesno',
                                            detail="'Sim' para Vertical, 'Não' para Horizontal")
            
            page_size = letter if orientacao == 'yes' else landscape(letter)
            
            escolha = messagebox.askquestion("Exportar", 
                                        "Deseja exportar a página atual ou todo o conteúdo?", 
                                        icon='question', type='yesnocancel',
                                        detail="'Sim' para página atual, 'Não' para todo conteúdo")
            
            if escolha == 'cancel':
                return
            
            dados_exportar = current_pagination.get_page() if escolha == 'yes' else dados
            
            caminho = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Salvar como PDF",
                initialfile=f"acervo_livros_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            )
            
            if not caminho:
                return

            try:
                # Configurações de layout
                margem = 40  # Margem uniforme
                espacamento_linha = 14
                altura_cabecalho = 20
                fonte_normal = 9
                fonte_cabecalho = 10
                fonte_titulo = 14
                
                # Larguras das colunas (ajustáveis automaticamente)
                larguras_colunas = {
                    "PRATELEIRA": 50,
                    "NUMERAÇÃO": 50,
                    "LIVROS": 140,
                    "AUTOR(ES)": 120,
                    "EDITORAS": 100,
                    "CLASSIFICAÇÃO/TEMA": 100,
                    "FAIXA ETÁRIA": 50
                }
                
                # Criar documento PDF
                doc = SimpleDocTemplate(caminho, pagesize=page_size,
                                    leftMargin=margem, rightMargin=margem,
                                    topMargin=margem, bottomMargin=margem)
                
                # Estilos
                estilos = getSampleStyleSheet()
                estilo_titulo = ParagraphStyle(
                    'Titulo',
                    parent=estilos['Title'],
                    fontSize=fonte_titulo,
                    alignment=TA_CENTER,
                    spaceAfter=20
                )
                
                estilo_cabecalho = ParagraphStyle(
                    'Cabecalho',
                    parent=estilos['Normal'],
                    fontSize=fonte_cabecalho,
                    textColor=colors.black,
                    alignment=TA_LEFT,
                    spaceAfter=10
                )
                
                estilo_texto = ParagraphStyle(
                    'Texto',
                    parent=estilos['Normal'],
                    fontSize=fonte_normal,
                    leading=12,
                    spaceAfter=5
                )
                
                # Conteúdo do PDF
                conteudo = []
                
                # Adicionar título
                conteudo.append(Paragraph(titulo, estilo_titulo))
                
                # Preparar dados para a tabela
                dados_tabela = []
                
                # Cabeçalho
                linha_cabecalho = []
                for campo in CAMPOS:
                    linha_cabecalho.append(Paragraph(f"<b>{campo}</b>", estilo_cabecalho))
                dados_tabela.append(linha_cabecalho)
                
                # Dados
                for row in dados_exportar:
                    linha = []
                    for i, valor in enumerate(row):
                        texto = str(valor) if valor is not None else ""
                        # Quebra de texto automática para células grandes
                        if len(texto) > 30 and CAMPOS[i] in ["LIVROS", "AUTOR(ES)", "EDITORAS", "CLASSIFICAÇÃO/TEMA"]:
                            texto = "<br/>".join([texto[j:j+30] for j in range(0, len(texto), 30)])
                        linha.append(Paragraph(texto, estilo_texto))
                    dados_tabela.append(linha)
                
                # Criar tabela
                tabela = Table(dados_tabela, colWidths=[larguras_colunas[campo] for campo in CAMPOS])
                
                # Estilo da tabela
                estilo_tabela = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), fonte_cabecalho),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ])
                
                # Alternar cores das linhas
                for i in range(1, len(dados_tabela)):
                    if i % 2 == 0:
                        estilo_tabela.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
                
                tabela.setStyle(estilo_tabela)
                conteudo.append(tabela)
                
                # Rodapé
                rodape = Paragraph(
                    f"<font size=8>Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | " +
                    f"Total de registros: {len(dados_exportar)}</font>",
                    estilos['Normal']
                )
                conteudo.append(Spacer(1, 12))
                conteudo.append(rodape)
                
                # Gerar PDF
                doc.build(conteudo)
                
                messagebox.showinfo(
                    "Exportado com sucesso",
                    f"PDF gerado com sucesso!\n\nLocal: {caminho}\n" +
                    f"Registros: {len(dados_exportar)}\n" +
                    f"Orientação: {'Vertical' if orientacao == 'yes' else 'Horizontal'}\n" +
                    f"Escopo: {'Página atual' if escolha == 'yes' else 'Todo o conteúdo'}",
                    parent=janela
                )
            
            except Exception as e:
                messagebox.showerror("Erro na exportação", 
                                f"Não foi possível gerar o PDF:\n{str(e)}", 
                                parent=janela)


        def exportar_opcao_csv():
            exportar_csv(filtered_data)
    
        def exportar_opcao_pdf():
            exportar_pdf(filtered_data)
    
        export_frame = tk.Frame(conteudo, bg="white")
        export_frame.pack(pady=5)
    
        tk.Label(export_frame, text="Exportar dados filtrados:", bg="white").pack(side="left", padx=5)
        tk.Button(export_frame, text="📤 CSV", command=exportar_opcao_csv, bg="#1996D4", fg="white").pack(side="left", padx=5)
        tk.Button(export_frame, text="📄 PDF", command=exportar_opcao_pdf, bg="#B36144", fg="white").pack(side="left", padx=5)

    def exibir_livros_doacao():
        limpar_conteudo()
        
        if not MEMORY_CACHE['ready']:
            carregando_frame = tk.Frame(conteudo, bg="white")
            carregando_frame.pack(expand=True, fill="both")
            tk.Label(carregando_frame, text="Carregando dados, aguarde...", font=("Arial", 14), bg="white").pack(expand=True)
            return
        
        dados_livros_doacao = MEMORY_CACHE['livros_doacao']
        paginacao = Paginacao(dados_livros_doacao)

        # Frame de filtro simplificado
        filtro_frame = tk.Frame(conteudo, bg="white")
        filtro_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(filtro_frame, text="🔍 Pesquisar (título, autor, editora, tema, etc.):", bg="white").pack(side="left")
        entrada_pesquisa = tk.Entry(filtro_frame, width=40)
        entrada_pesquisa.pack(side="left", padx=5)
        entrada_pesquisa.focus_set()  # Foco automático no campo de pesquisa

        # Botão de reiniciar/atualizar planilha
        def reiniciar_planilha():
            """Atualização rápida com feedback visual"""
            loading_label = tk.Label(conteudo, text="🔄 Atualizando...", font=("Arial", 12), bg="white")
            loading_label.pack(pady=10)
            janela.update()  # Atualiza a interface imediatamente
            
            def tarefa_atualizacao():
                try:
                    mod_time = os.path.getmtime(ARQUIVO_DOACAO)
                    
                    if mod_time > MEMORY_CACHE['last_update']:
                        dados = carregar_dados_livros_doacao()
                        MEMORY_CACHE['livros_doacao'] = dados
                        MEMORY_CACHE['last_update'] = time.time()
                        
                        janela.after(0, lambda: [
                            loading_label.destroy(),
                            exibir_livros_doacao(),
                            messagebox.showinfo("Sucesso", "Planilha de doações atualizada com sucesso!")
                        ])
                    else:
                        janela.after(0, lambda: [
                            loading_label.destroy(),
                            messagebox.showinfo("Info", "Nenhuma alteração encontrada na planilha de doações.")
                        ])
                        
                except Exception as e:
                    janela.after(0, lambda: [
                        loading_label.destroy(),
                        messagebox.showerror("Erro", f"Falha ao atualizar:\n{str(e)}")
                    ])
            
            threading.Thread(target=tarefa_atualizacao, daemon=True).start()
        
        tk.Button(filtro_frame, text="🔄 Atualizar Planilha", command=reiniciar_planilha, 
                bg="#16E127", fg="white").pack(side="right", padx=5)

        # Tabela
        tree_frame = tk.Frame(conteudo, bg="white")
        tree_frame.pack(fill="both", expand=True)

        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")

        # Usar lista para colunas
        tree = ttk.Treeview(tree_frame, columns=["#"] + CAMPOS, show="headings", 
                        yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        tree.pack(fill="both", expand=True)
        scroll_y.config(command=tree.yview)
        scroll_x.config(command=tree.xview)

        tree.heading("#", text="Nº")
        tree.column("#", width=50, anchor="center")

        for campo in CAMPOS:
            tree.heading(campo, text=campo)
            tree.column(campo, anchor="center", width=150)

        filtered_data = []
        current_pagination = paginacao

        def filtrar_livros():
            nonlocal filtered_data, current_pagination
            
            termo = entrada_pesquisa.get().lower()
            
            filtered_data = []
            
            for row in dados_livros_doacao:
                if not row or len(row) < 2:
                    continue
                    
                # Verifica em todos os campos relevantes
                campos_relevantes = [
                    str(row[0]) if len(row) > 0 else "",  # PRATELEIRA
                    str(row[1]) if len(row) > 1 else "",  # NUMERAÇÃO
                    str(row[2]) if len(row) > 2 else "",  # LIVROS (título)
                    str(row[3]) if len(row) > 3 else "",  # AUTOR(ES)
                    str(row[4]) if len(row) > 4 else "",  # EDITORAS
                    str(row[5]) if len(row) > 5 else "",  # CLASSIFICAÇÃO/TEMA
                    str(row[6]) if len(row) > 6 else ""   # FAIXA ETÁRIA
                ]
                
                # Verifica se o termo de pesquisa está em qualquer campo relevante
                if termo and any(termo in campo.lower() for campo in campos_relevantes):
                    filtered_data.append(row)
                elif not termo:  # Se não há termo de pesquisa, mostra tudo
                    filtered_data.append(row)
            
            current_pagination = Paginacao(filtered_data)
            update_table()

        def update_table():
            nonlocal current_pagination
            
            tree.delete(*tree.get_children())
            page_data = current_pagination.get_page()
            
            for i, row in enumerate(page_data, start=current_pagination.current_page * current_pagination.page_size):
                tree.insert("", "end", iid=i, values=(i+1,) + row)
            
            label_pagina.config(text=f"Página {current_pagination.current_page+1}/{current_pagination.total_pages}")
            botao_anterior.config(state="normal" if current_pagination.current_page > 0 else "disabled")
            botao_proximo.config(state="normal" if current_pagination.current_page < current_pagination.total_pages - 1 else "disabled")

        tk.Button(filtro_frame, text="🔎 Filtrar", command=filtrar_livros, bg="#2196F3", fg="white").pack(side="left", padx=10)

        # Controles de paginação
        paginacao_frame = tk.Frame(conteudo, bg="white")
        paginacao_frame.pack(pady=5)

        botao_anterior = tk.Button(paginacao_frame, text="◀ Anterior", command=lambda: [current_pagination.prev_page(), update_table()])
        botao_anterior.pack(side="left", padx=5)

        label_pagina = tk.Label(paginacao_frame, text=f"Página 1/{paginacao.total_pages}", bg="white")
        label_pagina.pack(side="left", padx=5)

        botao_proximo = tk.Button(paginacao_frame, text="Próximo ▶", command=lambda: [current_pagination.next_page(), update_table()])
        botao_proximo.pack(side="left", padx=5)

        # Preencher tabela inicial
        filtered_data = dados_livros_doacao
        update_table()

        # Botões de ação
        def get_selected_index():
            selecionado = tree.selection()
            if not selecionado:
                messagebox.showerror("Erro", "Selecione um livro.")
                return None
            return int(selecionado[0])

        acoes = tk.Frame(conteudo, bg="white")
        acoes.pack(pady=10)
        tk.Button(acoes, text="✏️ Atualizar", command=lambda: atualizar_livro_doacao(get_selected_index()), bg="#4CAF50", fg="white").pack(side="left", padx=10)
        tk.Button(acoes, text="🗑️ Excluir", 
                command=lambda: excluir_livros_doacao(get_selected_index()), 
                bg="#f44336", fg="white").pack(side="left", padx=10)



        # Exportação
        def exportar_csv(dados, titulo="ACERVO - LIVROS DE DOAÇÃO - BIBLIOTECA ANA ZÉLIA DE MORAIS LARA"):
            escolha = messagebox.askquestion(
                "Exportar", 
                "Deseja exportar a página atual ou todo o conteúdo?",
                icon='question', type='yesnocancel',
                detail="Selecione 'Sim' para página atual, 'Não' para todo conteúdo"
            )
            
            if escolha == 'cancel':
                return

            dados_exportar = current_pagination.get_page() if escolha == 'yes' else dados

            caminho = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Planilhas Excel", "*.xlsx")],
                title="Salvar como Excel",
                initialfile=f"acervo_livros_doacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            )

            if not caminho:
                return

            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Acervo Doação"

                # Estilos
                negrito = Font(bold=True)
                centralizado = Alignment(horizontal="center", vertical="center")
                borda_fina = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Título
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CAMPOS))
                cell_titulo = ws.cell(row=1, column=1)
                cell_titulo.value = titulo
                cell_titulo.font = Font(size=14, bold=True)
                cell_titulo.alignment = centralizado

                # Cabeçalho
                for col, campo in enumerate(CAMPOS, start=1):
                    cell = ws.cell(row=3, column=col, value=campo)
                    cell.font = negrito
                    cell.alignment = centralizado
                    cell.border = borda_fina

                # Dados
                for i, linha in enumerate(dados_exportar, start=4):
                    for j, valor in enumerate(linha, start=1):
                        cell = ws.cell(row=i, column=j, value=valor)
                        cell.alignment = Alignment(horizontal="left")
                        cell.border = borda_fina

                # Ajustar largura das colunas automaticamente
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column  # número da coluna
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ajuste = max_length + 2
                    ws.column_dimensions[get_column_letter(column)].width = ajuste

                # Linha de rodapé
                ws.append([])
                ws.append([f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
                ws.append([f"Total de registros: {len(dados_exportar)}"])

                # Salvar
                wb.save(caminho)

                messagebox.showinfo(
                    "Exportado com sucesso",
                    f"Planilha Excel gerada com sucesso!\n\nLocal: {caminho}\nRegistros exportados: {len(dados_exportar)}\nTipo: {'Página atual' if escolha == 'yes' else 'Todo o conteúdo'}",
                    parent=janela
                )

            except Exception as e:
                messagebox.showerror("Erro na exportação", f"Erro ao exportar os dados:\n{str(e)}", parent=janela)
            
        def exportar_pdf(dados, titulo="ACERVO - LIVROS DE DOAÇÃO - BIBLIOTECA ANA ZÉLIA DE MORAIS LARA"):
            orientacao = messagebox.askquestion("Orientação", 
                                            "Deseja página na vertical ou horizontal?", 
                                            icon='question', type='yesno',
                                            detail="'Sim' para Vertical, 'Não' para Horizontal")
            
            page_size = letter if orientacao == 'yes' else landscape(letter)
            
            escolha = messagebox.askquestion("Exportar", 
                                        "Deseja exportar a página atual ou todo o conteúdo?", 
                                        icon='question', type='yesnocancel',
                                        detail="'Sim' para página atual, 'Não' para todo conteúdo")
            
            if escolha == 'cancel':
                return
            
            dados_exportar = current_pagination.get_page() if escolha == 'yes' else dados
            
            caminho = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="Salvar como PDF",
                initialfile=f"acervo_livros_doacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            )
            
            if not caminho:
                return

            try:
                # Configurações de layout
                margem = 40  # Margem uniforme
                espacamento_linha = 14
                altura_cabecalho = 20
                fonte_normal = 9
                fonte_cabecalho = 10
                fonte_titulo = 14
                
                # Larguras das colunas (ajustáveis automaticamente)
                larguras_colunas = {
                    "PRATELEIRA": 50,
                    "NUMERAÇÃO": 50,
                    "LIVROS": 140,
                    "AUTOR(ES)": 120,
                    "EDITORAS": 100,
                    "CLASSIFICAÇÃO/TEMA": 100,
                    "FAIXA ETÁRIA": 50
                }
                
                # Criar documento PDF
                doc = SimpleDocTemplate(caminho, pagesize=page_size,
                                    leftMargin=margem, rightMargin=margem,
                                    topMargin=margem, bottomMargin=margem)
                
                # Estilos
                estilos = getSampleStyleSheet()
                estilo_titulo = ParagraphStyle(
                    'Titulo',
                    parent=estilos['Title'],
                    fontSize=fonte_titulo,
                    alignment=TA_CENTER,
                    spaceAfter=20
                )
                
                estilo_cabecalho = ParagraphStyle(
                    'Cabecalho',
                    parent=estilos['Normal'],
                    fontSize=fonte_cabecalho,
                    textColor=colors.black,
                    alignment=TA_LEFT,
                    spaceAfter=10
                )
                
                estilo_texto = ParagraphStyle(
                    'Texto',
                    parent=estilos['Normal'],
                    fontSize=fonte_normal,
                    leading=12,
                    spaceAfter=5
                )
                
                # Conteúdo do PDF
                conteudo = []
                
                # Adicionar título
                conteudo.append(Paragraph(titulo, estilo_titulo))
                
                # Preparar dados para a tabela
                dados_tabela = []
                
                # Cabeçalho
                linha_cabecalho = []
                for campo in CAMPOS:
                    linha_cabecalho.append(Paragraph(f"<b>{campo}</b>", estilo_cabecalho))
                dados_tabela.append(linha_cabecalho)
                
                # Dados
                for row in dados_exportar:
                    linha = []
                    for i, valor in enumerate(row):
                        texto = str(valor) if valor is not None else ""
                        # Quebra de texto automática para células grandes
                        if len(texto) > 30 and CAMPOS[i] in ["LIVROS", "AUTOR(ES)", "EDITORAS", "CLASSIFICAÇÃO/TEMA"]:
                            texto = "<br/>".join([texto[j:j+30] for j in range(0, len(texto), 30)])
                        linha.append(Paragraph(texto, estilo_texto))
                    dados_tabela.append(linha)
                
                # Criar tabela
                tabela = Table(dados_tabela, colWidths=[larguras_colunas[campo] for campo in CAMPOS])
                
                # Estilo da tabela
                estilo_tabela = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), fonte_cabecalho),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ])
                
                # Alternar cores das linhas
                for i in range(1, len(dados_tabela)):
                    if i % 2 == 0:
                        estilo_tabela.add('BACKGROUND', (0, i), (-1, i), colors.whitesmoke)
                
                tabela.setStyle(estilo_tabela)
                conteudo.append(tabela)
                
                # Rodapé
                rodape = Paragraph(
                    f"<font size=8>Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | " +
                    f"Total de registros: {len(dados_exportar)}</font>",
                    estilos['Normal']
                )
                conteudo.append(Spacer(1, 12))
                conteudo.append(rodape)
                
                # Gerar PDF
                doc.build(conteudo)
                
                messagebox.showinfo(
                    "Exportado com sucesso",
                    f"PDF gerado com sucesso!\n\nLocal: {caminho}\n" +
                    f"Registros: {len(dados_exportar)}\n" +
                    f"Orientação: {'Vertical' if orientacao == 'yes' else 'Horizontal'}\n" +
                    f"Escopo: {'Página atual' if escolha == 'yes' else 'Todo o conteúdo'}",
                    parent=janela
                )
            
            except Exception as e:
                messagebox.showerror("Erro na exportação", 
                                f"Não foi possível gerar o PDF:\n{str(e)}", 
                                parent=janela)


        def exportar_opcao_csv():
            exportar_csv(filtered_data)
    
        def exportar_opcao_pdf():
            exportar_pdf(filtered_data)
    
        export_frame = tk.Frame(conteudo, bg="white")
        export_frame.pack(pady=5)
    
        tk.Label(export_frame, text="Exportar dados filtrados:", bg="white").pack(side="left", padx=5)
        tk.Button(export_frame, text="📤 CSV", command=exportar_opcao_csv, bg="#1996D4", fg="white").pack(side="left", padx=5)
        tk.Button(export_frame, text="📄 PDF", command=exportar_opcao_pdf, bg="#B36144", fg="white").pack(side="left", padx=5)
    
    def modulo_emprestimos():
        limpar_conteudo()
        frame_emprestimos = tk.Frame(conteudo, bg="white")
        frame_emprestimos.pack(fill="both", expand=True)
    
        if not MEMORY_CACHE['ready']:
            tk.Label(frame_emprestimos, text="Carregando dados, aguarde...", font=("Arial", 14), bg="white").pack(expand=True)
            return
    
        dados_emprestimos = MEMORY_CACHE['emprestimos']
        dados_livros = MEMORY_CACHE['livros']
        paginacao = Paginacao(dados_emprestimos, 100)
    
        def buscar_dados_livro(codigo):
            for row in dados_livros:
                if row and len(row) > 1 and str(row[1]) == codigo:
                    return row[1], row[2], row[3]
            return None
    
        def livro_ja_emprestado(codigo):
            for row in dados_emprestimos:
                if row and len(row) > 0 and str(row[0]) == codigo:
                    return True
            return False
    
        def salvar_emprestimo():
            codigo = entrada_codigo.get().strip()
            if not codigo:
                messagebox.showerror("Erro", "Digite o código do livro")
                return
    
            if livro_ja_emprestado(codigo):
                messagebox.showwarning("Atenção", "Este livro já está emprestado. Aguarde a devolução para novo empréstimo.")
                return
    
            dados_livro = buscar_dados_livro(codigo)
            if not dados_livro:
                messagebox.showerror("Erro", "Livro não encontrado pelo código fornecido.")
                return
    
            _, titulo, autor = dados_livro
            aluno = entrada_aluno.get().strip()
            turma = entrada_turma.get().strip()
            try:
                prazo = int(entrada_prazo.get())
            except ValueError:
                messagebox.showerror("Erro", "Prazo deve ser um número inteiro")
                return
    
            data_hoje = datetime.today().strftime("%d/%m/%Y")
    
            if not aluno or not turma or prazo <= 0:
                messagebox.showerror("Erro", "Preencha todos os campos corretamente.")
                return
    
            try:
                wb = openpyxl.load_workbook(ARQUIVO_EMPRESTIMOS)
                ws = wb.active
                ws.append([codigo, titulo, autor, aluno, turma, prazo, data_hoje])
                wb.save(ARQUIVO_EMPRESTIMOS)
                
                threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
                messagebox.showinfo("Sucesso", "Empréstimo cadastrado com sucesso!")
                update_table()
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao salvar empréstimo: {str(e)}")
    
        def update_table(filtro_aluno="", filtro_turma=""):
            tree.delete(*tree.get_children())
            hoje = datetime.today()
            dados_filtrados = []
    
            for row in dados_emprestimos:
                if not row or len(row) < 6:
                    continue
                    
                codigo, titulo, autor, aluno, turma, prazo, data_emprestimo = row[:7]
                
                try:
                    if data_emprestimo is None or str(data_emprestimo).strip() == "":
                        data_emp = hoje
                        data_str = "Data não informada"
                    else:
                        data_emp = datetime.strptime(str(data_emprestimo), "%d/%m/%Y")
                        data_str = data_emprestimo
                except:
                    data_emp = hoje
                    data_str = "Data inválida"
                
                if filtro_aluno and filtro_aluno.lower() not in aluno.lower():
                    continue
                if filtro_turma and filtro_turma != turma and filtro_turma != "Todas":
                    continue
                
                nova_linha = (codigo, titulo, autor, aluno, turma, prazo, data_str)
                dados_filtrados.append(nova_linha)
                
                try:
                    prazo_int = int(prazo)
                    dias_passados = (hoje - data_emp).days
    
                    if dias_passados > prazo_int:
                        cor = "#df2323"
                    elif dias_passados > prazo_int // 2:
                        cor = "#faf60f"
                    else:
                        cor = "white"
                except:
                    cor = "white"
    
                tree.insert("", "end", values=nova_linha, tags=(cor,))
            
            for color in ["#df2323", "#faf60f", "white"]:
                tree.tag_configure(color, background=color)
            
            paginacao = Paginacao(dados_filtrados)
            label_pagina.config(text=f"Página {paginacao.current_page+1}/{paginacao.total_pages}")
            botao_anterior.config(state="normal" if paginacao.current_page > 0 else "disabled")
            botao_proximo.config(state="normal" if paginacao.current_page < paginacao.total_pages - 1 else "disabled")
    
        # Widgets de entrada
        frame_campos = tk.Frame(frame_emprestimos, bg="white")
        frame_campos.pack(pady=10)
    
        tk.Label(frame_campos, text="Código do Livro:", bg="white").grid(row=0, column=0, padx=5)
        entrada_codigo = tk.Entry(frame_campos)
        entrada_codigo.grid(row=0, column=1, padx=5)
        entrada_codigo.focus_set()  # Foco automático

        tk.Label(frame_campos, text="Aluno:", bg="white").grid(row=1, column=0, padx=5)
        entrada_aluno = tk.Entry(frame_campos)
        entrada_aluno.grid(row=1, column=1, padx=5)

        tk.Label(frame_campos, text="Turma:", bg="white").grid(row=2, column=0, padx=5)
        entrada_turma = tk.Entry(frame_campos)
        entrada_turma.grid(row=2, column=1, padx=5)

        tk.Label(frame_campos, text="Prazo (dias):", bg="white").grid(row=3, column=0, padx=5)
        entrada_prazo = tk.Entry(frame_campos)
        entrada_prazo.grid(row=3, column=1, padx=5)
        entrada_prazo.insert(0, "7")  # Valor padrão
    
        # Botões de ação
        frame_botoes = tk.Frame(frame_emprestimos, bg="white")
        frame_botoes.pack(pady=10)
        tk.Button(frame_botoes, text="Salvar Empréstimo", command=salvar_emprestimo, bg="#4CAF50", fg="white").pack(side="left", padx=5)
    
        # Filtros de pesquisa
        frame_filtro = tk.Frame(frame_emprestimos, bg="white")
        frame_filtro.pack(fill="x", padx=10, pady=10)
    
        tk.Label(frame_filtro, text="🔍 Pesquisar por aluno:", bg="white").pack(side="left")
        entrada_filtro_aluno = tk.Entry(frame_filtro, width=30)
        entrada_filtro_aluno.pack(side="left", padx=5)
    
        tk.Label(frame_filtro, text="🏫 Filtrar por turma:", bg="white").pack(side="left", padx=10)
        
        turmas = sorted(MEMORY_CACHE['turmas'])
        combo_filtro_turma = ttk.Combobox(frame_filtro, values=["Todas"] + turmas, state="readonly", width=15)
        combo_filtro_turma.current(0)
        combo_filtro_turma.pack(side="left", padx=5)
        
        # Botão de reiniciar/atualizar planilha
        def reiniciar_planilha():
            threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
            messagebox.showinfo("Atualização", "Dados da planilha serão atualizados. Aguarde alguns instantes.")
        
        tk.Button(frame_filtro, text="🔄 Atualizar Planilha", command=reiniciar_planilha, 
                 bg="#21F333", fg="white").pack(side="right", padx=5)
    
        def aplicar_filtros():
            aluno = entrada_filtro_aluno.get().strip()
            turma = combo_filtro_turma.get()
            update_table(aluno, turma)
    
        tk.Button(frame_filtro, text="🔎 Aplicar Filtros", command=aplicar_filtros, bg="#2196F3", fg="white").pack(side="left", padx=10)
    
        # Controles de paginação
        paginacao_frame = tk.Frame(frame_emprestimos, bg="white")
        paginacao_frame.pack(pady=5)
    
        botao_anterior = tk.Button(paginacao_frame, text="◀ Anterior", command=lambda: [paginacao.prev_page(), update_table()])
        botao_anterior.pack(side="left", padx=5)
    
        label_pagina = tk.Label(paginacao_frame, text="Página 1/1", bg="white")
        label_pagina.pack(side="left", padx=5)
    
        botao_proximo = tk.Button(paginacao_frame, text="Próximo ▶", command=lambda: [paginacao.next_page(), update_table()])
        botao_proximo.pack(side="left", padx=5)
    
        # Tabela
        frame_tabela = tk.Frame(frame_emprestimos, bg="white")
        frame_tabela.pack(fill="both", expand=True, padx=10, pady=10)
    
        tree = ttk.Treeview(frame_tabela, columns=CAMPOS_EMPRESTIMO, show="headings")
        vsb = ttk.Scrollbar(frame_tabela, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
    
        for col in CAMPOS_EMPRESTIMO:
            tree.heading(col, text=col)
            tree.column(col, width=100)
    
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
    
        for color in ["#ffcccc", "#fff2cc", "white"]:
            tree.tag_configure(color, background=color)
    
        update_table()
    
    def exibir_emprestimos_ativos():
        limpar_conteudo()
        frame = tk.Frame(conteudo, bg="white")
        frame.pack(fill="both", expand=True)
    
        if not MEMORY_CACHE['ready']:
            tk.Label(frame, text="Carregando dados, aguarde...", font=("Arial", 14), bg="white").pack(expand=True)
            return
    
        # Título
        tk.Label(
            frame,
            text="📚 LIVROS EMPRESTADOS NO MOMENTO",
            font=("Segoe UI", 14, "bold"),
            bg="white",
            pady=10
        ).pack(fill="x")
    
        # Frame de filtros
        filtro_frame = tk.Frame(frame, bg="white")
        filtro_frame.pack(fill="x", padx=10, pady=5)
    
        tk.Label(filtro_frame, text="🔍 Buscar por aluno:", bg="white").pack(side="left")
        entrada_aluno = tk.Entry(filtro_frame)
        entrada_aluno.pack(side="left", padx=5)
    
        tk.Label(filtro_frame, text="🏫 Turma:", bg="white").pack(side="left", padx=10)
        
        turmas = sorted({e[4] for e in MEMORY_CACHE['emprestimos'] if len(e) > 4 and e[4]})
        combo_turma = ttk.Combobox(filtro_frame, values=["Todas"] + list(turmas), state="readonly", width=15)
        combo_turma.current(0)
        combo_turma.pack(side="left", padx=5)
        
        # Botão de reiniciar/atualizar planilha
        def reiniciar_planilha():
            threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
            messagebox.showinfo("Atualização", "Dados da planilha serão atualizados. Aguarde alguns instantes.")
        
        tk.Button(filtro_frame, text="🔄 Atualizar Planilha", command=reiniciar_planilha, 
                 bg="#16E127", fg="white").pack(side="right", padx=5)
    
        tk.Button(filtro_frame, text="Aplicar Filtros", command=lambda: atualizar_tabela(), 
                 bg="#16E127", fg="white").pack(side="left", padx=10)
    
        # Frame da tabela
        tabela_frame = tk.Frame(frame, bg="white")
        tabela_frame.pack(fill="both", expand=True, padx=10, pady=5)
    
        scroll_y = ttk.Scrollbar(tabela_frame)
        scroll_x = ttk.Scrollbar(tabela_frame, orient="horizontal")
    
        tree = ttk.Treeview(tabela_frame, columns=CAMPOS_EMPRESTIMO, show="headings",
                           yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        scroll_y.config(command=tree.yview)
        scroll_x.config(command=tree.xview)
    
        for col in CAMPOS_EMPRESTIMO:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor="center")
    
        tree.pack(side="left", fill="both", expand=True)
        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")
    
        # Frame de botões de ação
        botoes_frame = tk.Frame(frame, bg="white")
        botoes_frame.pack(pady=10)
    
        tk.Button(botoes_frame, text="✏️ Editar Prazo", command=lambda: editar_prazo(), 
                 bg="#FFC107", fg="black").pack(side="left", padx=5)
    
        tk.Button(botoes_frame, text="🗑️ Devolver/Excluir", command=lambda: devolver_livro(), 
                 bg="#f44336", fg="white").pack(side="left", padx=5)
    
        def atualizar_tabela():
            tree.delete(*tree.get_children())
            filtro_aluno = entrada_aluno.get().lower()
            filtro_turma = combo_turma.get()
            hoje = datetime.today()
            
            for row in MEMORY_CACHE['emprestimos']:
                if len(row) < 7:
                    continue
                    
                codigo, titulo, autor, aluno, turma, prazo, data_emp = row[:7]
                
                if filtro_aluno and filtro_aluno not in aluno.lower():
                    continue
                if filtro_turma != "Todas" and filtro_turma != turma:
                    continue
                
                try:
                    data_emprestimo = datetime.strptime(data_emp, "%d/%m/%Y")
                    dias_passados = (hoje - data_emprestimo).days
                    prazo_int = int(prazo)
                    
                    if dias_passados > prazo_int:
                        cor = "#cf2323"
                    elif dias_passados > prazo_int // 2:
                        cor = "#ddc918"
                    else:
                        cor = "white"
                except:
                    cor = "white"
                
                tree.insert("", "end", values=row, tags=(cor,))
                tree.tag_configure(cor, background=cor)
    
        def editar_prazo():
            selecionado = tree.selection()
            if not selecionado:
                messagebox.showwarning("Aviso", "Selecione um empréstimo para editar")
                return
                
            item = tree.item(selecionado[0])
            valores = item['values']
            
            novo_prazo = simpledialog.askinteger("Editar Prazo", 
                                               f"Digite o novo prazo (dias) para:\n\nLivro: {valores[1]}\nAluno: {valores[3]}",
                                               parent=janela,
                                               minvalue=1)
            
            if novo_prazo:
                try:
                    wb = openpyxl.load_workbook(ARQUIVO_EMPRESTIMOS)
                    ws = wb.active
                    
                    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        if (str(row[0].value) == str(valores[0]) and
                            str(row[3].value) == str(valores[3]) and
                            str(row[6].value) == str(valores[6])):
                            
                            ws.cell(row=idx, column=6, value=novo_prazo)
                            break
                    
                    wb.save(ARQUIVO_EMPRESTIMOS)
                    
                    threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
                    messagebox.showinfo("Sucesso", "Prazo atualizado com sucesso!")
                    atualizar_tabela()
                    
                except Exception as e:
                    messagebox.showerror("Erro", f"Falha ao atualizar prazo:\n{str(e)}")

        def devolver_livro():
            selecionado = tree.selection()
            if not selecionado:
                messagebox.showwarning("Aviso", "Selecione um empréstimo para devolver")
                return
                
            item = tree.item(selecionado[0])
            valores = item['values']
            
            senha = simpledialog.askstring("Confirmação", 
                                        "Digite a senha para confirmar a devolução:",
                                        show="*",
                                        parent=janela)
            
            if senha != "123":
                messagebox.showerror("Erro", "Senha incorreta!")
                return
                
            if not messagebox.askyesno("Confirmar", 
                                    f"Confirmar devolução do livro:\n\nTítulo: {valores[1]}\nAluno: {valores[3]}\nData: {valores[6]}?"):
                return
            
            try:
                wb = openpyxl.load_workbook(ARQUIVO_EMPRESTIMOS)
                ws = wb.active
                
                # Procurar o registro correto
                encontrado = False
                registro_apagado = None
                for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                    
                    if len(row_values) < 7:
                        continue
                    
                    # CORREÇÃO AQUI - Parênteses corretamente fechados
                    if (row_values[0] == str(valores[0]) and 
                        row_values[3] == str(valores[3]) and 
                        row_values[6] == str(valores[6])):
                        
                        registro_apagado = {
                            'codigo': row[0].value,
                            'titulo': row[1].value,
                            'autor': row[2].value,
                            'aluno': row[3].value,
                            'turma': row[4].value,
                            'data_devolucao': datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                        }
                        ws.delete_rows(idx)
                        encontrado = True
                        break
                
                if not encontrado:
                    messagebox.showerror("Erro", "Registro não encontrado na planilha!")
                    return
                
                wb.save(ARQUIVO_EMPRESTIMOS)
                
                # Salvar no histórico
                if registro_apagado:
                    try:
                        historico_path = "historico_emprestimos.xlsx"
                        
                        if os.path.exists(historico_path):
                            wb_historico = openpyxl.load_workbook(historico_path)
                            ws_historico = wb_historico["Histórico"]
                        else:
                            wb_historico = openpyxl.Workbook()
                            ws_historico = wb_historico.active
                            ws_historico.title = "Histórico"
                            ws_historico.append(["CÓDIGO", "Título do Livro", "Autor", "Nome do Aluno", "turma", "Data de DEVOLUÇÃO"])
                        
                        ws_historico.append([
                            registro_apagado['codigo'],
                            registro_apagado['titulo'],
                            registro_apagado['autor'],
                            registro_apagado['aluno'],
                            registro_apagado['turma'],
                            registro_apagado['data_devolucao']
                        ])
                        
                        wb_historico.save(historico_path)
                        
                    except Exception as e_historico:
                        messagebox.showwarning("Aviso", 
                                            f"Devolução registrada, mas falha ao salvar no histórico:\n{str(e_historico)}")
                        with open("erro_historico.log", "a") as f:
                            f.write(f"{datetime.now()} - Erro ao salvar histórico: {str(e_historico)}\n")
                
                threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
                messagebox.showinfo("Sucesso", "Livro devolvido com sucesso!")
                atualizar_tabela()
                
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao registrar devolução:\n{str(e)}")


        atualizar_tabela()
    
    # Botões do menu
    tk.Button(menu_content, text="➕ Novo livro", command=novo_livro, **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="📖 Exibir todos os livros", command=exibir_livros, **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="➕ Novo livro doação", command=novo_livro_doacao, **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="📦 Exibir todos os livros doação", command=exibir_livros_doacao, **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="🤵 Novo Aluno", command=cadastrar_aluno, **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="📚 Realizar um empréstimo", command=modulo_emprestimos, **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="📌 Livros emprestados no momento", command=exibir_emprestimos_ativos, **estilo_botao).pack(pady=2)
    
    # Função para voltar ao login
    def voltar_login():
        if messagebox.askyesno("Confirmação", "Deseja realmente sair e voltar à tela de login?"):
            janela.destroy()
            root = tk.Tk()
            app = LoginApp(root)
            root.mainloop()
    
    # Botão de logout no canto inferior esquerdo
    logout_btn = tk.Button(
        menu_bottom,
        text="🚪 Sair",
        command=voltar_login,
        bg="#444444",
        fg="white",
        font=("Segoe UI", 10, "bold"),
        padx=10,
        pady=8,
        relief="flat",
        cursor="hand2"
    )
    logout_btn.pack(side="left", padx=10, pady=10, fill="x", expand=True)
    
    # Sistema de atualização assíncrona
    def verificar_carregamento():
        try:
            while not DATA_QUEUE.empty():
                msg, livros_count, doacao_count, emprestimos_count = DATA_QUEUE.get_nowait()
                if msg == 'dados_carregados':
                    carregando_frame.destroy()
                    exibir_livros()
        except queue.Empty:
            pass
        
        if not MEMORY_CACHE['ready']:
            janela.after(500, verificar_carregamento)
    
    threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
    janela.after(500, verificar_carregamento)
    

    
    janela.mainloop()





def iniciar_sistema_aluno(matricula):
    """Inicia o sistema para o aluno com as funcionalidades específicas"""
    criar_planilha_if_not_exists()
    
    janela = tk.Tk()
    janela.title(f"Sistema de Biblioteca - Aluno {ALUNOS[matricula]['nome']}")
    janela.geometry("1100x600")
    janela.configure(bg="#2c2c2c")

    # === Controle de inatividade ===
    janela.last_activity = time.time()
    
    def reset_inactivity_timer(event=None):
        janela.last_activity = time.time()
    
    def check_inactivity():
        if time.time() - janela.last_activity > 1800:  # 30 minutos
            if messagebox.askyesno("Sessão Expirada", 
                                  "Sua sessão expirou por inatividade. Deseja continuar?"):
                reset_inactivity_timer()
            else:
                janela.destroy()
                root = tk.Tk()
                app = LoginApp(root)
                root.mainloop()
        else:
            janela.after(60000, check_inactivity)  # Verificar a cada minuto
    
    # Registrar eventos para resetar o timer
    janela.bind("<Key>", reset_inactivity_timer)
    janela.bind("<Button>", reset_inactivity_timer)
    janela.after(60000, check_inactivity)  # Iniciar verificação
    
    # Componentes da interface
    topo = tk.Frame(janela, bg="#3498db", height=70)
    topo.pack(side="top", fill="x")
    
    titulo = tk.Label(
        topo,
        text=f"📖 BIBLIOTECA SMART | ALUNO: {ALUNOS[matricula]['nome']} - {ALUNOS[matricula]['turma']}",
        bg="#3498db",
        fg="white",
        font=("Segoe UI", 16, "bold")
    )
    titulo.pack(pady=10, expand=True)
    
    menu = tk.Frame(janela, bg="#1c1c1c", width=240)
    menu.pack(side="left", fill="y")
    
    # Criar container para botões do menu
    menu_content = tk.Frame(menu, bg="#1c1c1c")
    menu_content.pack(side="top", fill="both", expand=True)
    
    # Criar frame para botão de logout no rodapé
    menu_bottom = tk.Frame(menu, bg="#1c1c1c")
    menu_bottom.pack(side="bottom", fill="x")
    
    conteudo = tk.Frame(janela, bg="white")
    conteudo.pack(expand=True, fill="both")
    
    carregando_frame = tk.Frame(conteudo, bg="white")
    carregando_frame.pack(expand=True, fill="both")
    tk.Label(carregando_frame, text="Carregando dados, aguarde...", font=("Arial", 14), bg="white").pack(expand=True)
    
    estilo_botao = {
        "bg": "#333333",
        "fg": "white",
        "font": ("Segoe UI", 10),
        "relief": "flat",
        "width": 30,
        "anchor": "w",
        "padx": 10,
        "pady": 5
    }
    
    def limpar_conteudo():
        for widget in conteudo.winfo_children():
            widget.destroy()
    
    def exibir_emprestimos_atuais(matricula):
        """Mostra todos os livros que o aluno pegou e ainda não devolveu"""
        limpar_conteudo()
        
        if not MEMORY_CACHE['ready']:
            tk.Label(conteudo, text="Carregando dados, aguarde...", font=("Arial", 14), bg="white").pack(expand=True)
            return
        
        aluno_info = ALUNOS[matricula]
        dados_emprestimos = MEMORY_CACHE['emprestimos']
        
        # Filtrar empréstimos atuais do aluno
        emprestimos_atuais = []
        hoje = datetime.today()
        
        nome_completo_aluno = aluno_info['nome'].strip().lower()
        
        for emp in dados_emprestimos:
            if len(emp) < 7:
                continue
                
            nome_emp = str(emp[3]).strip().lower() if emp[3] is not None else ""
            
            if nome_emp == nome_completo_aluno:
                try:
                    data_emprestimo_str = str(emp[6]) if emp[6] is not None else ""
                    data_emp = hoje
                    
                    if data_emprestimo_str:
                        try:
                            data_emp = datetime.strptime(data_emprestimo_str, "%d/%m/%Y")
                        except ValueError:
                            data_emp = hoje
                    
                    prazo = 7
                    if emp[5] is not None:
                        try:
                            prazo = int(emp[5])
                        except (ValueError, TypeError):
                            print(f"Prazo inválido: '{emp[5]}', usando 7 dias")
                    
                    data_devolucao = data_emp + timedelta(days=prazo)
                    status = "Em atraso" if hoje > data_devolucao else "No prazo"
                    
                    emprestimos_atuais.append({
                        "id": emp[0],
                        "titulo": emp[1],
                        "autor": emp[2],
                        "data_emprestimo": data_emprestimo_str,
                        "data_devolucao": data_devolucao.strftime("%d/%m/%Y"),
                        "status": status,
                        "turma": emp[4]
                    })
                except Exception as e:
                    print(f"Erro ao processar empréstimo: {e}")
        
        # Frame principal com design moderno
        frame_principal = tk.Frame(conteudo, bg="#f0f8ff")
        frame_principal.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        tk.Label(
            frame_principal,
            text="📚 MEUS EMPRÉSTIMOS ATUAIS",
            font=("Segoe UI", 16, "bold"),
            bg="#f0f8ff",
            fg="#2c3e50"
        ).pack(pady=(0, 20))
        
        if not emprestimos_atuais:
            tk.Label(
                frame_principal,
                text="Você não tem nenhum livro emprestado no momento.",
                font=("Segoe UI", 12),
                bg="#f0f8ff",
                fg="#3498db"
            ).pack(expand=True)
            return
        
        # Container com sombra visual
        container_tabela = tk.Frame(frame_principal, bg="#e0e7f0", padx=1, pady=1)
        container_tabela.pack(fill="both", expand=True)
        
        # Frame branco da tabela
        frame_tabela = tk.Frame(container_tabela, bg="white", padx=5, pady=5)
        frame_tabela.pack(fill="both", expand=True, padx=1, pady=1)
        
        # Estilo Treeview
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#e0e0e0", foreground="black")
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=25)
        style.map("Treeview", background=[('selected', '#2980b9')])
        
        # Treeview com ID oculto
        tree = ttk.Treeview(
            frame_tabela,
            columns=("id", "Título", "Autor", "Data Empréstimo", "Data Devolução", "Status"),
            show="headings",
            height=min(10, len(emprestimos_atuais))
        )
        
        # Cabeçalhos centralizados
        tree.heading("id", text="ID", anchor="center")
        tree.heading("Título", text="TÍTULO", anchor="center")
        tree.heading("Autor", text="AUTOR", anchor="center")
        tree.heading("Data Empréstimo", text="DATA EMPRÉSTIMO", anchor="center")
        tree.heading("Data Devolução", text="DATA DEVOLUÇÃO", anchor="center")
        tree.heading("Status", text="STATUS", anchor="center")
        
        # Colunas centralizadas
        tree.column("id", width=0, stretch=tk.NO)
        tree.column("Título", width=200, anchor="center")
        tree.column("Autor", width=150, anchor="center")
        tree.column("Data Empréstimo", width=120, anchor="center")
        tree.column("Data Devolução", width=120, anchor="center")
        tree.column("Status", width=100, anchor="center")
        
        # Scrollbars
        vsb = ttk.Scrollbar(frame_tabela, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame_tabela, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        frame_tabela.grid_rowconfigure(0, weight=1)
        frame_tabela.grid_columnconfigure(0, weight=1)
        
        # Inserir dados na tabela com cor por status
        for emp in emprestimos_atuais:
            tag = 'atraso' if emp["status"] == "Em atraso" else 'prazo'
            tree.insert("", "end", values=(
                emp["id"],
                emp["titulo"],
                emp["autor"],
                emp["data_emprestimo"],
                emp["data_devolucao"],
                emp["status"]
            ), tags=(tag,))
        
        # Cores alternadas
        tree.tag_configure("atraso", background="#ffcccc")
        tree.tag_configure("prazo", background="#e6ffe6")
        
        # Rodapé com informações
        rodape = tk.Frame(frame_principal, bg="#f0f8ff")
        rodape.pack(fill="x", pady=(20, 0))
        
        tk.Label(
            rodape,
            text=f"Total de livros emprestados: {len(emprestimos_atuais)}",
            font=("Segoe UI", 10, "bold"),
            bg="#f0f8ff",
            fg="#2c3e50"
        ).pack(side="left")
        
        data_atualizacao = datetime.now().strftime("%d/%m/%Y %H:%M")
        tk.Label(
            rodape,
            text=f"Atualizado em: {data_atualizacao}",
            font=("Segoe UI", 9),
            bg="#f0f8ff",
            fg="#7f8c8d"
        ).pack(side="right")

        


    def exibir_historico_emprestimos(matricula):
        """Exibe todos os livros que o aluno já emprestou no passado a partir do histórico"""
        limpar_conteudo()
        
        aluno_info = ALUNOS[matricula]
        nome_aluno = aluno_info['nome'].strip().lower()
        
        # Frame principal com gradiente suave
        frame_principal = tk.Frame(conteudo, bg="#f0f8ff")  # Azul claro suave
        frame_principal.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título com estilo moderno
        titulo_frame = tk.Frame(frame_principal, bg="#f0f8ff")
        titulo_frame.pack(fill="x", pady=(0, 20))
        
        tk.Label(
            titulo_frame,
            text="📚 HISTÓRICO DE EMPRÉSTIMOS",
            font=("Segoe UI", 16, "bold"),
            bg="#f0f8ff",
            fg="#2c3e50"
        ).pack(pady=(10, 5))
        
        tk.Label(
            titulo_frame,
            text=f"Aluno: {aluno_info['nome']} | Turma: {aluno_info['turma']}",
            font=("Segoe UI", 12),
            bg="#f0f8ff",
            fg="#7f8c8d"
        ).pack()
        
        # Container para a tabela com sombra visual
        container_tabela = tk.Frame(frame_principal, bg="#e0e7f0", padx=1, pady=1)
        container_tabela.pack(fill="both", expand=True)
        
        # Frame da tabela com fundo branco
        frame_tabela = tk.Frame(container_tabela, bg="white", padx=5, pady=5)
        frame_tabela.pack(fill="both", expand=True, padx=1, pady=1)
        
        try:
            # Carregar o histórico de empréstimos
            historico_path = "historico_emprestimos.xlsx"
            if not os.path.exists(historico_path):
                mensagem = tk.Label(
                    frame_principal,
                    text="Histórico de empréstimos não encontrado.",
                    font=("Segoe UI", 12),
                    bg="#f0f8ff",
                    fg="#e74c3c"
                )
                mensagem.pack(expand=True, pady=20)
                return
                
            wb = openpyxl.load_workbook(historico_path)
            ws = wb["Histórico"]
            
            historico = []
            cabecalhos = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            
            # Encontrar índices das colunas
            colunas = {
                'codigo': None,
                'titulo': None,
                'autor': None,
                'aluno': None,
                'turma': None,
                'data_devolucao': None
            }
            
            for idx, col in enumerate(cabecalhos):
                col_str = str(col).strip()
                if col_str == "CÓDIGO":
                    colunas['codigo'] = idx
                elif col_str == "Título do Livro":
                    colunas['titulo'] = idx
                elif col_str == "Autor":
                    colunas['autor'] = idx
                elif col_str == "Nome do Aluno":
                    colunas['aluno'] = idx
                elif col_str == "turma":
                    colunas['turma'] = idx
                elif col_str == "Data de DEVOLUÇÃO":
                    colunas['data_devolucao'] = idx
            
            # Verificar se encontramos todas as colunas necessárias
            if None in colunas.values():
                messagebox.showerror("Erro", "O arquivo de histórico não tem a estrutura esperada!")
                wb.close()
                return
            
            # Percorrer as linhas
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Verificar se a linha tem dados suficientes
                if len(row) < max(colunas.values())+1:
                    continue
                    
                nome_historico = str(row[colunas['aluno']]).strip().lower() if row[colunas['aluno']] else ""
                
                if nome_historico == nome_aluno:
                    # Formatar a data de devolução
                    data_devolucao = row[colunas['data_devolucao']]
                    if isinstance(data_devolucao, datetime):
                        data_devolucao_str = data_devolucao.strftime("%d/%m/%Y %H:%M")
                    else:
                        data_devolucao_str = str(data_devolucao) if data_devolucao else "Não registrada"
                    
                    historico.append({
                        "titulo": row[colunas['titulo']],
                        "autor": row[colunas['autor']],
                        "data_devolucao": data_devolucao_str
                    })
            
            wb.close()
            
            if not historico:
                mensagem = tk.Label(
                    frame_principal,
                    text="Você ainda não tem histórico de empréstimos registrados.",
                    font=("Segoe UI", 12),
                    bg="#f8f9fa",
                    fg="#3498db"
                )
                mensagem.pack(expand=True, pady=20)
                return
            
            # Configurar Treeview com estilo moderno
            style = ttk.Style()
            style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#3498db", foreground="black")
            style.configure("Treeview", font=("Segoe UI", 10), rowheight=25)
            style.map("Treeview", background=[('selected', "#2980b9")])
            
            # Configurar Treeview
            tree = ttk.Treeview(
                frame_tabela,
                columns=("Título", "Autor", "Data Devolução"),
                show="headings",
                height=min(10, len(historico))  # Ajusta altura conforme número de itens
            )
            
            # Configurar cabeçalhos
            tree.heading("Título", text="TÍTULO DO LIVRO", anchor="center")
            tree.heading("Autor", text="AUTOR", anchor="center")
            tree.heading("Data Devolução", text="DATA DE DEVOLUÇÃO", anchor="center")
            
            # Configurar colunas
            tree.column("Título", width=300, anchor="center")
            tree.column("Autor", width=250, anchor="center")
            tree.column("Data Devolução", width=180, anchor="center")
            
            # Adicionar scrollbars
            vsb = ttk.Scrollbar(frame_tabela, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(frame_tabela, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            # Posicionar widgets
            tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            
            # Configurar grid
            frame_tabela.grid_rowconfigure(0, weight=1)
            frame_tabela.grid_columnconfigure(0, weight=1)
            
            # Adicionar dados à tabela com cores alternadas
            for i, item in enumerate(historico):
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                tree.insert("", "end", values=(
                    item["titulo"],
                    item["autor"],
                    item["data_devolucao"]
                ), tags=(tag,))
            
            # Configurar tags para cores alternadas
            tree.tag_configure('evenrow', background='#f8f9fa')
            tree.tag_configure('oddrow', background='#ffffff')
            
            # Rodapé com informações
            rodape = tk.Frame(frame_principal, bg="#f0f8ff")
            rodape.pack(fill="x", pady=(20, 0))
            
            tk.Label(
                rodape,
                text=f"Total de empréstimos: {len(historico)}",
                font=("Segoe UI", 10, "bold"),
                bg="#f0f8ff",
                fg="#2c3e50"
            ).pack(side="left")
            
            # Adicionar data de atualização
            data_atualizacao = datetime.now().strftime("%d/%m/%Y %H:%M")
            tk.Label(
                rodape,
                text=f"Atualizado em: {data_atualizacao}",
                font=("Segoe UI", 9),
                bg="#f0f8ff",
                fg="#7f8c8d"
            ).pack(side="right")
            
            # Função para abrir janela de edição
            def abrir_janela_edicao(titulo_livro, autor_livro, comentario_existente="", nota_existente=0):
                janela = tk.Toplevel()
                janela.title("⭐ AVALIAÇÃO DE LIVRO")
                janela.geometry("400x600")
                janela.resizable(False, False)
                janela.configure(bg="#f0f8ff")
                janela.transient(janela.master)
                janela.grab_set()
                
                # Frame principal com padding
                frame_principal = tk.Frame(janela, bg="#f0f8ff", padx=20, pady=20)
                frame_principal.pack(fill="both", expand=True)
                
                # Cabeçalho com estilo consistente
                cabecalho_frame = tk.Frame(frame_principal, bg="#f0f8ff")
                cabecalho_frame.pack(fill="x", pady=(0, 15))
                
                tk.Label(
                    cabecalho_frame,
                    text="⭐ AVALIAR LIVRO",
                    font=("Segoe UI", 16, "bold"),
                    bg="#f0f8ff",
                    fg="#2c3e50"
                ).pack(pady=(0, 5))
                
                # Card do livro com sombra visual
                card_livro = tk.Frame(frame_principal, bg="white", padx=15, pady=12, relief="groove", bd=1)
                card_livro.pack(fill="x", pady=(0, 20))
                
                tk.Label(
                    card_livro, 
                    text=titulo_livro,
                    font=("Segoe UI", 12, "bold"), 
                    bg="white",
                    wraplength=350,
                    justify="left"
                ).pack(anchor="w", pady=(0, 5))
                
                tk.Label(
                    card_livro, 
                    text=f"por {autor_livro}", 
                    font=("Segoe UI", 10), 
                    bg="white",
                    fg="#7f8c8d"
                ).pack(anchor="w", pady=(0, 10))
                
                tk.Label(
                    card_livro, 
                    text=f"Aluno: {aluno_info['nome']}", 
                    font=("Segoe UI", 10), 
                    bg="white"
                ).pack(anchor="w")
                
                tk.Label(
                    card_livro, 
                    text=f"Turma: {aluno_info['turma']}", 
                    font=("Segoe UI", 10), 
                    bg="white"
                ).pack(anchor="w")
                
                # Seção de avaliação
                avaliacao_frame = tk.Frame(frame_principal, bg="#f0f8ff")
                avaliacao_frame.pack(fill="x", pady=(0, 15))
                
                tk.Label(
                    avaliacao_frame,
                    text="Sua Avaliação:",
                    font=("Segoe UI", 12, "bold"),
                    bg="#f0f8ff",
                    fg="#2c3e50"
                ).pack(anchor="w", pady=(0, 10))
                
                # Avaliação por estrelas (5 estrelas)
                estrelas_frame = tk.Frame(avaliacao_frame, bg="#f0f8ff")
                estrelas_frame.pack()
                
                nota_var = tk.IntVar(value=nota_existente)
                estrelas = []
                
                def atualizar_estrelas(n):
                    nota_var.set(n)
                    for i, estrela in enumerate(estrelas):
                        estrela.config(text="★" if i < n else "☆", fg="#f39c12" if i < n else "#bdc3c7")
                
                for i in range(5):
                    estrela = tk.Label(
                        estrelas_frame, 
                        text="★" if i < nota_existente else "☆", 
                        font=("Segoe UI", 28), 
                        bg="#f0f8ff", 
                        cursor="hand2",
                        fg="#f39c12" if i < nota_existente else "#bdc3c7"
                    )
                    estrela.pack(side="left", padx=3)
                    estrela.bind("<Enter>", lambda e, n=i+1: atualizar_estrelas(n) if nota_var.get() == 0 else None)
                    estrela.bind("<Leave>", lambda e: atualizar_estrelas(nota_var.get()))
                    estrela.bind("<Button-1>", lambda e, n=i+1: [atualizar_estrelas(n), janela.focus_set()])
                    estrelas.append(estrela)
                
                # Comentário com ScrolledText
                comentario_frame = tk.Frame(frame_principal, bg="#f0f8ff")
                comentario_frame.pack(fill="x", pady=(15, 0))
                
                tk.Label(
                    comentario_frame,
                    text="Comentário (opcional):",
                    font=("Segoe UI", 10, "bold"),
                    bg="#f0f8ff"
                ).pack(anchor="w", pady=(0, 5))
                
                campo_comentario = scrolledtext.ScrolledText(
                    comentario_frame,
                    height=6,
                    width=40,
                    wrap="word",
                    font=("Segoe UI", 10),
                    padx=10,
                    pady=10,
                    bg="white",
                    bd=2,
                    relief="sunken"
                )
                campo_comentario.pack(fill="x")
                if comentario_existente:
                    campo_comentario.insert("1.0", comentario_existente)
                
                # Botões com estilo moderno
                botoes_frame = tk.Frame(frame_principal, bg="#f0f8ff")
                botoes_frame.pack(fill="x", pady=(20, 0), padx=10)
                
                btn_container = tk.Frame(botoes_frame, bg="#f0f8ff")
                btn_container.pack(expand=True)
                
                def salvar_avaliacao():
                    nota = nota_var.get()
                    comentario = campo_comentario.get("1.0", "end-1c").strip()
                    
                    if nota == 0:
                        messagebox.showwarning("Aviso", "Por favor, avalie o livro com 1 a 5 estrelas.", parent=janela)
                        return
                    
                    try:
                        planilha = "avaliacoes.xlsx"
                        caminho_completo = os.path.abspath(planilha)
                        
                        nova_avaliacao = {
                            "Título": titulo_livro,
                            "Autor": autor_livro,
                            "Aluno": aluno_info["nome"],
                            "Turma": aluno_info["turma"],
                            "Comentário": comentario,
                            "Nota": nota,
                            "Data": datetime.now().strftime("%d/%m/%Y %H:%M")
                        }
                        
                        if os.path.exists(planilha):
                            df = pd.read_excel(planilha)
                            if comentario_existente:
                                # Atualizar avaliação existente
                                idx = df[
                                    (df['Título'] == titulo_livro) &
                                    (df['Aluno'] == aluno_info['nome'])
                                ].index
                                if not idx.empty:
                                    df.loc[idx, ['Comentário', 'Nota', 'Data']] = [comentario, nota, nova_avaliacao['Data']]
                            else:
                                df = pd.concat([df, pd.DataFrame([nova_avaliacao])], ignore_index=True)
                        else:
                            df = pd.DataFrame([nova_avaliacao])
                        
                        df.to_excel(planilha, index=False, engine='openpyxl')
                        
                        messagebox.showinfo(
                            "Avaliação Salva", 
                            f"Sua avaliação foi registrada com sucesso!\n\nArquivo salvo em:\n{caminho_completo}",
                            parent=janela
                        )
                        janela.destroy()
                        
                    except Exception as e:
                        messagebox.showerror("Erro", f"Não foi possível salvar a avaliação:\n{str(e)}", parent=janela)
                
                btn_salvar = tk.Button(
                    btn_container,
                    text="💾 SALVAR AVALIAÇÃO",
                    font=("Segoe UI", 10, "bold"),
                    bg="#27ae60",
                    fg="white",
                    padx=20,
                    pady=5,
                    command=salvar_avaliacao
                )
                btn_salvar.pack(side="right", ipadx=10)
                
                btn_cancelar = tk.Button(
                    btn_container,
                    text="CANCELAR",
                    font=("Segoe UI", 10),
                    bg="#e74c3c",
                    fg="white",
                    padx=20,
                    pady=5,
                    command=janela.destroy
                )
                btn_cancelar.pack(side="right", padx=10, ipadx=10)
                
                # Centralizar janela
                janela.update_idletasks()
                width = janela.winfo_width()
                height = janela.winfo_height()
                x = (janela.winfo_screenwidth() // 2) - (width // 2)
                y = (janela.winfo_screenheight() // 2) - (height // 2)
                janela.geometry(f'{width}x{height}+{x}+{y}')
            
            # Botão AVALIAR LIVRO
            def abrir_janela_avaliacao():
                selecionado = tree.focus()
                if not selecionado:
                    messagebox.showwarning("Aviso", "Selecione um livro na tabela para avaliar.")
                    return
                
                dados = tree.item(selecionado)["values"]
                if len(dados) < 2:
                    messagebox.showerror("Erro", "Dados do livro incompletos.")
                    return
                
                titulo_livro = dados[0]
                autor_livro = dados[1]
                
                # Verificar se o aluno já avaliou o livro
                if os.path.exists("avaliacoes.xlsx"):
                    df_avaliacoes = pd.read_excel("avaliacoes.xlsx")
                    avaliacao_existente = df_avaliacoes[
                        (df_avaliacoes['Título'] == titulo_livro) &
                        (df_avaliacoes['Aluno'] == aluno_info['nome']) &
                        (df_avaliacoes['Comentário'] != '')
                    ]
                    
                    if not avaliacao_existente.empty:
                        resposta = messagebox.askyesno(
                            "Comentário Existente",
                            "Você já fez um comentário para este livro. Deseja visualizar/editar seu comentário?"
                        )
                        if resposta:
                            # Abrir janela de visualização/edição do comentário
                            janela_comentario = tk.Toplevel()
                            janela_comentario.title(f"Comentário: {titulo_livro}")
                            janela_comentario.geometry("350x550")
                            janela_comentario.resizable(False, False)
                            janela_comentario.configure(bg="#f8fafc")
                            janela_comentario.transient(janela_comentario.master)
                            janela_comentario.grab_set()
                            
                            # Cabeçalho
                            header_frame = tk.Frame(janela_comentario, bg="#4b6cb7", height=70)
                            header_frame.pack(fill="x")
                            
                            tk.Label(
                                header_frame,
                                text="💬 Seu Comentário",
                                font=("Segoe UI", 14, "bold"),
                                bg="#4b6cb7",
                                fg="white",
                                pady=20
                            ).pack()
                            
                            tk.Label(
                                header_frame,
                                text=f"Livro: {titulo_livro}",
                                font=("Segoe UI", 10),
                                bg="#4b6cb7",
                                fg="#e2e8f0"
                            ).pack()
                            
                            # Frame principal com scroll
                            main_frame = tk.Frame(janela_comentario, bg="#f8fafc")
                            main_frame.pack(fill="both", expand=True, padx=10, pady=10)
                            
                            # Card do comentário
                            comentario = avaliacao_existente.iloc[0]
                            card = tk.Frame(
                                main_frame,
                                bg="white",
                                bd=0,
                                relief="solid",
                                highlightbackground="#e2e8f0",
                                highlightthickness=1,
                                padx=15,
                                pady=12
                            )
                            card.pack(fill="x", pady=8, padx=5)
                            
                            header = tk.Frame(card, bg="white")
                            header.pack(fill="x", pady=(0, 8))
                            
                            tk.Label(
                                header,
                                text=f"{comentario['Aluno']}",
                                font=("Segoe UI", 10, "bold"),
                                bg="white",
                                anchor="w"
                            ).pack(side="left")
                            
                            stars = "★" * int(comentario['Nota'])
                            tk.Label(
                                header,
                                text=stars,
                                font=("Segoe UI", 10),
                                fg="#f59e0b",
                                bg="white",
                                anchor="w"
                            ).pack(side="left", padx=5)
                            
                            if 'Data' in comentario and pd.notna(comentario['Data']):
                                if isinstance(comentario['Data'], str):
                                    data_str = comentario['Data']
                                else:
                                    data_str = comentario['Data'].strftime("%d/%m/%Y %H:%M")
                                
                                tk.Label(
                                    header,
                                    text=f"· {data_str}",
                                    font=("Segoe UI", 8),
                                    fg="#64748b",
                                    bg="white",
                                    anchor="e"
                                ).pack(side="right")
                            
                            tk.Label(
                                card,
                                text=comentario['Comentário'],
                                font=("Segoe UI", 10),
                                bg="white",
                                wraplength=300,
                                justify="left",
                                anchor="w"
                            ).pack(fill="x")
                            
                            # Botões de edição e exclusão
                            botoes_frame = tk.Frame(janela_comentario, bg="#f8fafc")
                            botoes_frame.pack(fill="x", pady=10, padx=10)
                            
                            def excluir_comentario():
                                if messagebox.askyesno("Confirmação", "Deseja realmente excluir este comentário?", parent=janela_comentario):
                                    try:
                                        df_avaliacoes.drop(avaliacao_existente.index, inplace=True)
                                        df_avaliacoes.to_excel("avaliacoes.xlsx", index=False, engine='openpyxl')
                                        messagebox.showinfo("Sucesso", "Comentário excluído com sucesso!", parent=janela_comentario)
                                        janela_comentario.destroy()
                                    except Exception as e:
                                        messagebox.showerror("Erro", f"Não foi possível excluir o comentário:\n{str(e)}", parent=janela_comentario)
                            
                            def editar_comentario():
                                janela_comentario.destroy()
                                abrir_janela_edicao(titulo_livro, autor_livro, comentario['Comentário'], comentario['Nota'])
                            
                            tk.Button(
                                botoes_frame,
                                text="✕ Excluir Comentário",
                                font=("Segoe UI", 10),
                                bg="#e74c3c",
                                fg="white",
                                bd=1,
                                relief="flat",
                                highlightbackground="#e2e8f0",
                                highlightthickness=1,
                                command=excluir_comentario
                            ).pack(side="left", padx=5, pady=5, ipadx=10)
                            
                            tk.Button(
                                botoes_frame,
                                text="✎ Editar Comentário",
                                font=("Segoe UI", 10),
                                bg="#3498db",
                                fg="white",
                                bd=1,
                                relief="flat",
                                highlightbackground="#e2e8f0",
                                highlightthickness=1,
                                command=editar_comentario
                            ).pack(side="left", padx=5, pady=5, ipadx=10)
                            
                            # Centralizar janela
                            janela_comentario.update_idletasks()
                            width = janela_comentario.winfo_width()
                            height = janela_comentario.winfo_height()
                            x = (janela_comentario.winfo_screenwidth() // 2) - (width // 2)
                            y = (janela_comentario.winfo_screenheight() // 2) - (height // 2)
                            janela_comentario.geometry(f'{width}x{height}+{x}+{y}')
                            
                            return
                
                # Chama a janela de edição diretamente se não houver comentário ou se o aluno não quiser visualizar
                abrir_janela_edicao(titulo_livro, autor_livro)
            
            # Botão final abaixo da tabela
            btn_avaliar = tk.Button(
                frame_principal,
                text="📖 Avaliar Livro Selecionado",
                font=("Segoe UI", 10, "bold"),
                bg="#f39c12",
                fg="white",
                command=abrir_janela_avaliacao
            )
            btn_avaliar.pack(pady=15)
            
        except Exception as e:
            # Tratamento de erro com estilo
            erro_frame = tk.Frame(frame_principal, bg="#f0f8ff")
            erro_frame.pack(expand=True, pady=20)
            
            tk.Label(
                erro_frame,
                text="⚠️ ERRO AO CARREGAR HISTÓRICO",
                font=("Segoe UI", 12, "bold"),
                bg="#f0f8ff",
                fg="#e74c3c"
            ).pack(pady=(0, 10))
            
            tk.Label(
                erro_frame,
                text=str(e),
                font=("Segoe UI", 10),
                bg="#f0f8ff",
                fg="#2c3e50",
                wraplength=500
            ).pack()
            
            tk.Label(
                erro_frame,
                text="Por favor, tente novamente mais tarde ou entre em contato com o suporte.",
                font=("Segoe UI", 10),
                bg="#f0f8ff",
                fg="#7f8c8d"
            ).pack(pady=(10, 0))



    def exibir_livros_alunos(matricula):
        limpar_conteudo()

        caminho_planilha = "LIVROS LITERÁRIOS  -  ALUNOS.xlsx"
        if not os.path.exists(caminho_planilha):
            tk.Label(conteudo, text="Arquivo de livros não encontrado.", font=("Arial", 14), bg="white").pack(expand=True)
            return

        try:
            df = pd.read_excel(caminho_planilha)
            dados_livros = df.values.tolist()
        except Exception as e:
            tk.Label(conteudo, text=f"Erro ao ler a planilha: {str(e)}", font=("Arial", 12), fg="red", bg="white").pack(expand=True)
            return

        current_page = 0
        page_size = 100
        filtered_data = list(dados_livros)

        # ==== FUNDO E CABEÇALHO MODERNO ====
        frame_principal = tk.Frame(conteudo, bg="#f0f8ff")  # Azul claro suave
        frame_principal.pack(fill="both", expand=True, padx=20, pady=20)

        titulo_frame = tk.Frame(frame_principal, bg="#f0f8ff")
        titulo_frame.pack(fill="x", pady=(0, 20))

        tk.Label(
            titulo_frame,
            text="📚 LIVROS DA BIBLIOTECA",
            font=("Segoe UI", 16, "bold"),
            bg="#f0f8ff",
            fg="#2c3e50"
        ).pack(pady=(10, 5))

        # ==== FILTRO ====
        filtro_frame = tk.Frame(frame_principal, bg="#f0f8ff")
        filtro_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(filtro_frame, text="🔍 Pesquisar (título, autor, editora, tema, etc.):", bg="#f0f8ff").pack(side="left")
        entrada_pesquisa = tk.Entry(filtro_frame, width=40)
        entrada_pesquisa.pack(side="left", padx=5)
        entrada_pesquisa.focus_set()

        tk.Button(filtro_frame, text="🔎 Filtrar", command=lambda: filtrar_livros(), bg="#2196F3", fg="white").pack(side="left", padx=10)

        # ==== TABELA ====
        tree_frame = tk.Frame(frame_principal, bg="white")
        tree_frame.pack(fill="both", expand=True)

        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")

        colunas = list(df.columns)
        tree = ttk.Treeview(tree_frame, columns=["#"] + colunas, show="headings",
                            yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        tree.pack(fill="both", expand=True)
        scroll_y.config(command=tree.yview)
        scroll_x.config(command=tree.xview)

        tree.heading("#", text="Nº", anchor="center")
        tree.column("#", width=50, anchor="center")

        for campo in colunas:
            tree.heading(campo, text=campo, anchor="center")
            tree.column(campo, anchor="center", width=150)

        # ==== PAGINAÇÃO ====
        paginacao_frame = tk.Frame(frame_principal, bg="#f0f8ff")
        paginacao_frame.pack(pady=5)

        botao_anterior = tk.Button(paginacao_frame, text="◀ Anterior", command=lambda: anterior())
        botao_anterior.pack(side="left", padx=5)

        label_pagina = tk.Label(paginacao_frame, text=f"Página 1/{(len(filtered_data) + page_size - 1) // page_size}", bg="#f0f8ff")
        label_pagina.pack(side="left", padx=5)

        botao_proximo = tk.Button(paginacao_frame, text="Próximo ▶", command=lambda: proximo())
        botao_proximo.pack(side="left", padx=5)

        # ==== FUNÇÕES INTERNAS ====
        def get_total_pages():
            return max(1, (len(filtered_data) + page_size - 1) // page_size)

        def get_page_data():
            start = current_page * page_size
            end = start + page_size
            return filtered_data[start:end]

        def update_table():
            nonlocal current_page
            tree.delete(*tree.get_children())
            page_data = get_page_data()

            for i, row in enumerate(page_data, start=current_page * page_size):
                tree.insert("", "end", iid=i, values=[i + 1] + list(row))

            total_pages = get_total_pages()
            label_pagina.config(text=f"Página {current_page + 1}/{total_pages}")
            botao_anterior.config(state="normal" if current_page > 0 else "disabled")
            botao_proximo.config(state="normal" if current_page < total_pages - 1 else "disabled")

        def filtrar_livros():
            nonlocal filtered_data, current_page
            termo = entrada_pesquisa.get().lower()
            filtered_data = []

            for row in dados_livros:
                texto_linha = " ".join(str(campo).lower() for campo in row)
                if termo in texto_linha:
                    filtered_data.append(row)
                elif not termo:
                    filtered_data.append(row)

            current_page = 0
            update_table()

        def anterior():
            nonlocal current_page
            if current_page > 0:
                current_page -= 1
                update_table()

        def proximo():
            nonlocal current_page
            if current_page < get_total_pages() - 1:
                current_page += 1
                update_table()

        update_table()



    
    def exibir_ranking_alunos():
        """Mostra os alunos que mais pegaram livros emprestados a partir da planilha"""
        limpar_conteudo()



        arquivo = "historico_emprestimos.xlsx"
        if not os.path.exists(arquivo):
            tk.Label(conteudo, text="Arquivo de histórico não encontrado.", font=("Arial", 14), bg="white").pack(expand=True)
            return

        # Carrega planilha
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active

        contador = defaultdict(int)
        turmas_alunos = {}
        cabecalhos = [cell.value for cell in ws[1]]

        try:
            idx_nome = cabecalhos.index("Nome do Aluno")
            idx_turma = cabecalhos.index("turma")
        except ValueError:
            tk.Label(conteudo, text="Colunas 'Nome do Aluno' ou 'turma' não encontradas.", font=("Arial", 14), bg="white").pack(expand=True)
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            nome = str(row[idx_nome]).strip() if row[idx_nome] else ""
            turma = str(row[idx_turma]).strip() if row[idx_turma] else ""
            if nome and turma:
                chave = (nome, turma)
                contador[chave] += 1
                turmas_alunos[chave] = turma
        wb.close()

        ranking = sorted(contador.items(), key=lambda x: x[1], reverse=True)
        ranking = [
            {"aluno": nome, "turma": turma, "total": total}
            for (nome, turma), total in ranking
        ]

        # Interface (mesmo estilo moderno usado antes)
        frame_principal = tk.Frame(conteudo, bg="#f0f8ff")
        frame_principal.pack(fill="both", expand=True, padx=20, pady=20)

        tk.Label(frame_principal, text="🏆 RANKING DE LEITORES",
                font=("Segoe UI", 16, "bold"), bg="#f0f8ff", fg="#2c3e50").pack(pady=(0, 20))

        frame_filtros = tk.Frame(frame_principal, bg="#f0f8ff")
        frame_filtros.pack(fill="x", pady=(0, 10))

        tk.Label(frame_filtros, text="Filtrar por turma:", bg="#f0f8ff", font=("Segoe UI", 10)).pack(side="left")
        turmas = sorted({a["turma"] for a in ranking})
        combo_turma = ttk.Combobox(frame_filtros, values=["Todas"] + turmas, state="readonly")
        combo_turma.current(0)
        combo_turma.pack(side="left", padx=5)

        tk.Label(frame_filtros, text="Filtrar por nome:", bg="#f0f8ff", font=("Segoe UI", 10)).pack(side="left", padx=(20, 5))
        entrada_nome = tk.Entry(frame_filtros)
        entrada_nome.pack(side="left", padx=5)

        def atualizar_ranking():
            turma_filtro = combo_turma.get().lower()
            nome_filtro = entrada_nome.get().strip().lower()

            for i in tree.get_children():
                tree.delete(i)

            pos = 1
            for aluno in ranking:
                nome = aluno["aluno"].strip().lower()
                turma = aluno["turma"].strip().lower()
                if (turma_filtro == "todas" or turma == turma_filtro) and (nome_filtro in nome):
                    tree.insert("", "end", values=(
                        f"{pos}º",
                        aluno["aluno"],
                        aluno["turma"],
                        aluno["total"]
                    ))
                    pos += 1

        tk.Button(frame_filtros, text="🔍 Aplicar Filtro", command=atualizar_ranking,
                bg="#3498db", fg="white", font=("Segoe UI", 10, "bold")).pack(side="left", padx=10)

        container_tabela = tk.Frame(frame_principal, bg="#e0e7f0", padx=1, pady=1)
        container_tabela.pack(fill="both", expand=True)

        frame_tabela = tk.Frame(container_tabela, bg="white", padx=5, pady=5)
        frame_tabela.pack(fill="both", expand=True)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#e0e0e0", foreground="black")
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=25)
        style.map("Treeview", background=[('selected', '#2980b9')])

        tree = ttk.Treeview(frame_tabela, columns=("Posição", "Aluno", "Turma", "Total de Livros"),
                            show="headings", height=10)
        tree.heading("Posição", text="POSIÇÃO", anchor="center")
        tree.heading("Aluno", text="ALUNO", anchor="center")
        tree.heading("Turma", text="TURMA", anchor="center")
        tree.heading("Total de Livros", text="TOTAL DE LIVROS", anchor="center")

        tree.column("Posição", width=80, anchor="center")
        tree.column("Aluno", width=200, anchor="center")
        tree.column("Turma", width=100, anchor="center")
        tree.column("Total de Livros", width=120, anchor="center")

        vsb = ttk.Scrollbar(frame_tabela, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame_tabela, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        frame_tabela.grid_rowconfigure(0, weight=1)
        frame_tabela.grid_columnconfigure(0, weight=1)

        atualizar_ranking()

        rodape = tk.Frame(frame_principal, bg="#f0f8ff")
        rodape.pack(fill="x", pady=(20, 0))

        tk.Label(rodape, text="Exibindo os alunos com mais empréstimos registrados.",
                font=("Segoe UI", 10), bg="#f0f8ff", fg="#2c3e50").pack(side="left")

        tk.Label(rodape, text=f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                font=("Segoe UI", 9), bg="#f0f8ff", fg="#7f8c8d").pack(side="right")


    
    def exibir_livros_avaliados():
        """Exibe os livros com informações de empréstimos e avaliações"""
        limpar_conteudo()
        
        # Frame principal com o mesmo estilo das outras funções
        frame_principal = tk.Frame(conteudo, bg="#f0f8ff")
        frame_principal.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título com estilo consistente
        tk.Label(
            frame_principal,
            text="📊 LIVROS COM AVALIAÇÕES",
            font=("Segoe UI", 16, "bold"),
            bg="#f0f8ff",
            fg="#2c3e50"
        ).pack(pady=(0, 20))
        
        try:
            # Carregar dados
            historico_path = "historico_emprestimos.xlsx"
            avaliacoes_path = "avaliacoes.xlsx"
            
            if not os.path.exists(historico_path):
                raise FileNotFoundError("Arquivo de histórico não encontrado")
            
            # Processar dados de empréstimos
            df_historico = pd.read_excel(historico_path)
            df_historico = df_historico.dropna(subset=["Título do Livro"])
            
            emprestimos = df_historico.groupby("Título do Livro").size().reset_index(name="Empréstimos")
            emprestimos.columns = ["Título", "Empréstimos"]
            
            # Processar avaliações se o arquivo existir
            df_avaliacoes = pd.DataFrame()
            if os.path.exists(avaliacoes_path):
                df_avaliacoes = pd.read_excel(avaliacoes_path)
                df_avaliacoes['Nota'] = pd.to_numeric(df_avaliacoes['Nota'], errors='coerce').fillna(0)
                df_avaliacoes['Comentário'] = df_avaliacoes['Comentário'].fillna('')
                
                avaliacoes = df_avaliacoes.groupby(['Título', 'Autor']).agg({
                    'Nota': 'mean',
                    'Comentário': lambda x: sum(x != '')
                }).reset_index()
                avaliacoes.columns = ['Título', 'Autor', 'Média', 'Comentários']
                avaliacoes['Média'] = avaliacoes['Média'].round(1)
            else:
                avaliacoes = pd.DataFrame(columns=['Título', 'Autor', 'Média', 'Comentários'])
            
            # Combinar os dados
            livros = pd.merge(
                emprestimos,
                df_historico[['Título do Livro', 'Autor']].drop_duplicates(),
                left_on='Título',
                right_on='Título do Livro',
                how='left'
            ).drop(columns=['Título do Livro'])
            
            livros = pd.merge(livros, avaliacoes, on=['Título', 'Autor'], how='left')
            livros['Média'] = livros['Média'].fillna(0)
            livros['Comentários'] = livros['Comentários'].fillna(0)
            livros = livros.astype({'Média': 'float64', 'Comentários': 'int64'})
            livros = livros.sort_values(by='Empréstimos', ascending=False)
            
            # Container da tabela com sombra visual
            container_tabela = tk.Frame(frame_principal, bg="#e0e7f0", padx=1, pady=1)
            container_tabela.pack(fill="both", expand=True)
            
            # Frame branco da tabela
            frame_tabela = tk.Frame(container_tabela, bg="white", padx=5, pady=5)
            frame_tabela.pack(fill="both", expand=True, padx=1, pady=1)
            
            # Configurar estilo da Treeview
            style = ttk.Style()
            style.configure("Treeview.Heading", 
                        font=("Segoe UI", 10, "bold"), 
                        background="#e0e0e0", 
                        foreground="black")
            style.configure("Treeview", 
                        font=("Segoe UI", 10), 
                        rowheight=25,
                        fieldbackground="white")
            style.map("Treeview", background=[('selected', '#2980b9')])
            
            # Criar Treeview
            tree = ttk.Treeview(
                frame_tabela,
                columns=("Título", "Autor", "Empréstimos", "Avaliação", "Comentários"),
                show="headings",
                height=min(12, len(livros))
            )
            
            # Configurar cabeçalhos
            tree.heading("Título", text="TÍTULO", anchor="center")
            tree.heading("Autor", text="AUTOR", anchor="center")
            tree.heading("Empréstimos", text="EMPRÉSTIMOS", anchor="center")
            tree.heading("Avaliação", text="AVALIAÇÃO", anchor="center")
            tree.heading("Comentários", text="COMENTÁRIOS", anchor="center")
            
            # Configurar colunas
            tree.column("Título", width=250, anchor="center")
            tree.column("Autor", width=200, anchor="center")
            tree.column("Empréstimos", width=100, anchor="center")
            tree.column("Avaliação", width=120, anchor="center")
            tree.column("Comentários", width=120, anchor="center")
            
            # Adicionar scrollbars
            vsb = ttk.Scrollbar(frame_tabela, orient="vertical", command=tree.yview)
            hsb = ttk.Scrollbar(frame_tabela, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            # Posicionar widgets
            tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            
            # Configurar grid
            frame_tabela.grid_rowconfigure(0, weight=1)
            frame_tabela.grid_columnconfigure(0, weight=1)
            
            # Adicionar dados à tabela com cores alternadas
            for i, (_, row) in enumerate(livros.iterrows()):
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                avaliacao = f"{row['Média']:.1f} ★" if row['Média'] > 0 else "Sem avaliações"
                tree.insert("", "end", values=(
                    row['Título'],
                    row['Autor'],
                    row['Empréstimos'],
                    avaliacao,
                    row['Comentários']
                ), tags=(tag,))
            
            # Configurar tags para cores alternadas
            tree.tag_configure('evenrow', background='#f8f9fa')
            tree.tag_configure('oddrow', background='#ffffff')

            # Botão de comentários
            def exibir_comentarios():
                selecionado = tree.focus()
                if not selecionado:
                    messagebox.showwarning("Aviso", "Selecione um livro para ver os comentários")
                    return

                dados = tree.item(selecionado)["values"]
                titulo_livro = dados[0]

                if not os.path.exists("avaliacoes.xlsx"):
                    messagebox.showinfo("Informação", "Nenhuma avaliação registrada ainda")
                    return

                comentarios = df_avaliacoes[(df_avaliacoes['Título'] == titulo_livro) &
                                            (df_avaliacoes['Comentário'] != '')]

                if comentarios.empty:
                    messagebox.showinfo("Informação", f"Não há comentários para '{titulo_livro}'")
                    return

                # Janela
                janela_comentarios = tk.Toplevel()
                janela_comentarios.title(f"Comentários: {titulo_livro}")
                janela_comentarios.geometry("350x550")
                janela_comentarios.resizable(False, False)
                janela_comentarios.configure(bg="#f8fafc")

                # Cabeçalho
                header_frame = tk.Frame(janela_comentarios, bg="#4b6cb7", height=70)
                header_frame.pack(fill="x")

                tk.Label(
                    header_frame,
                    text="💬 Comentários",
                    font=("Segoe UI", 14, "bold"),
                    bg="#4b6cb7",
                    fg="white",
                    pady=20
                ).pack()

                tk.Label(
                    header_frame,
                    text=f"Livro: {titulo_livro}",
                    font=("Segoe UI", 10),
                    bg="#4b6cb7",
                    fg="#e2e8f0"
                ).pack()

                # Frame principal com scroll
                main_frame = tk.Frame(janela_comentarios, bg="#f8fafc")
                main_frame.pack(fill="both", expand=True, padx=10, pady=10)

                canvas = tk.Canvas(main_frame, bg="#f8fafc", bd=0, highlightthickness=0)
                scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
                scrollable_frame = tk.Frame(canvas, bg="#f8fafc")

                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                )

                canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", width=570)
                canvas.configure(yscrollcommand=scrollbar.set)

                canvas.pack(side="left", fill="both", expand=True)
                scrollbar.pack(side="right", fill="y")

                # Comentários em cards mais largos
                for _, comentario in comentarios.iterrows():
                    card = tk.Frame(
                        scrollable_frame,
                        bg="white",
                        bd=0,
                        relief="solid",
                        highlightbackground="#e2e8f0",
                        highlightthickness=1,
                        padx=15,
                        pady=12
                    )
                    card.pack(fill="x", pady=8, padx=5)

                    header = tk.Frame(card, bg="white")
                    header.pack(fill="x", pady=(0, 8))

                    tk.Label(
                        header,
                        text=f"{comentario['Aluno']}",
                        font=("Segoe UI", 10, "bold"),
                        bg="white",
                        anchor="w"
                    ).pack(side="left")

                    stars = "★" * int(comentario['Nota'])
                    tk.Label(
                        header,
                        text=stars,
                        font=("Segoe UI", 10),
                        fg="#f59e0b",
                        bg="white",
                        anchor="w"
                    ).pack(side="left", padx=5)

                    if 'Data' in comentario and pd.notna(comentario['Data']):
                        if isinstance(comentario['Data'], str):
                            data_str = comentario['Data']
                        else:
                            data_str = comentario['Data'].strftime("%d/%m/%Y %H:%M")

                        tk.Label(
                            header,
                            text=f"· {data_str}",
                            font=("Segoe UI", 8),
                            fg="#64748b",
                            bg="white",
                            anchor="e"
                        ).pack(side="right")

                    # Comentário maior, com largura ajustada
                    tk.Label(
                        card,
                        text=comentario['Comentário'],
                        font=("Segoe UI", 10),
                        bg="white",
                        wraplength=520,
                        justify="left",
                        anchor="w"
                    ).pack(fill="x")

                # Botão de fechar
                tk.Button(
                    janela_comentarios,
                    text="✕ Fechar",
                    font=("Segoe UI", 10),
                    bg="white",
                    fg="#64748b",
                    bd=1,
                    relief="flat",
                    highlightbackground="#e2e8f0",
                    highlightthickness=1,
                    command=janela_comentarios.destroy
                ).pack(pady=10, ipadx=20)


                

            
            # Botão moderno
            tk.Button(
                frame_principal,
                text="💬 Ver Comentários",
                font=("Segoe UI", 11, "bold"),
                bg="#3498db",
                fg="white",
                activebackground="#2980b9",
                activeforeground="white",
                bd=0,
                relief="flat",
                command=exibir_comentarios
            ).pack(pady=15)

            rodape = tk.Frame(frame_principal, bg="#f0f8ff")
            rodape.pack(fill="x", pady=(20, 0))

            tk.Label(
                rodape,
                text="Dados combinados de empréstimos e avaliações",
                font=("Segoe UI", 10),
                bg="#f0f8ff",
                fg="#2c3e50"
            ).pack(side="left")

            data_atualizacao = datetime.now().strftime("%d/%m/%Y %H:%M")
            tk.Label(
                rodape,
                text=f"Atualizado em: {data_atualizacao}",
                font=("Segoe UI", 9),
                bg="#f0f8ff",
                fg="#7f8c8d"
            ).pack(side="right")

        except Exception as e:
            erro = tk.Frame(frame_principal, bg="#f0f8ff")
            erro.pack(expand=True, pady=20)

            tk.Label(
                erro,
                text="⚠️ ERRO AO CARREGAR DADOS",
                font=("Segoe UI", 12, "bold"),
                bg="#f0f8ff",
                fg="#e74c3c"
            ).pack(pady=(0, 10))

            tk.Label(
                erro,
                text=str(e),
                font=("Segoe UI", 10),
                bg="#f0f8ff",
                fg="#2c3e50",
                wraplength=500
            ).pack()

            tk.Label(
                erro,
                text="Verifique se os arquivos existem e estão no formato correto.",
                font=("Segoe UI", 10),
                bg="#f0f8ff",
                fg="#7f8c8d"
            ).pack(pady=(10, 0))



    
    # Botões do menu para aluno
    tk.Button(menu_content, text="📚 Meus Empréstimos Atuais", 
              command=lambda: exibir_emprestimos_atuais(matricula), **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="📖 Meu Histórico de Empréstimos", 
              command=lambda: exibir_historico_emprestimos(matricula), **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="📕 Livros da Biblioteca", 
            command=lambda: exibir_livros_alunos(matricula), **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="🏆 Ranking de Leitores", 
              command=exibir_ranking_alunos, **estilo_botao).pack(pady=2)
    tk.Button(menu_content, text="📊 Avaliações de Livros", 
              command=exibir_livros_avaliados, **estilo_botao).pack(pady=2)
    

    
    # Função para voltar ao login
    def voltar_login():
        if messagebox.askyesno("Confirmação", "Deseja realmente sair e voltar à tela de login?"):
            janela.destroy()
            root = tk.Tk()
            app = LoginApp(root)
            root.mainloop()
    
    # Botão de logout no canto inferior esquerdo
    logout_btn = tk.Button(
        menu_bottom,
        text="🚪 Sair",
        command=voltar_login,
        bg="#444444",
        fg="white",
        font=("Segoe UI", 10, "bold"),
        padx=10,
        pady=8,
        relief="flat",
        cursor="hand2"
    )
    logout_btn.pack(side="left", padx=10, pady=10, fill="x", expand=True)
    
    # Sistema de atualização assíncrona
    def verificar_carregamento():
        try:
            while not DATA_QUEUE.empty():
                msg, livros_count, doacao_count, emprestimos_count = DATA_QUEUE.get_nowait()
                if msg == 'dados_carregados':
                    carregando_frame.destroy()
                    exibir_emprestimos_atuais(matricula)
        except queue.Empty:
            pass
        
        if not MEMORY_CACHE['ready']:
            janela.after(500, verificar_carregamento)
    
    threading.Thread(target=carregar_dados_em_segundo_plano, daemon=True).start()
    janela.after(500, verificar_carregamento)
    
    # Exibir empréstimos atuais por padrão
    janela.after(1000, lambda: exibir_emprestimos_atuais(matricula))


    janela.mainloop()



# Inicialização do sistema
if __name__ == "__main__":
    root = tk.Tk()
    app = LoginApp(root)
    root.mainloop()